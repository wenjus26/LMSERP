from functools import wraps
from io import BytesIO
import os
import secrets
from datetime import datetime
from weasyprint import HTML

import pandas as pd
from flask import (
    Flask, abort, flash,  make_response, render_template,
    request, redirect, session, url_for, send_file
)
from flask_login import (
    LoginManager, login_user, logout_user, login_required, 
    current_user 
)
from psycopg2 import IntegrityError
import pytz
from sqlalchemy import case, func
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

from models import (
    LogAction, Role, SalesLoadingInfo, SalesTruckDriverInfo, SalesWeightInfo,
    User, db, Customer, Contract, PaymentInfo, DateInfo, RemarksInfo, 
    FreightForwarder, Booking, Container, Loading, Seal, Weight
)
from validators import (
    validate_email, validate_full_name, validate_container_number,
    validate_truck_number, validate_phone_number,
    validate_number, validate_password, validate_contract_number
)


# Configuration de l'application Flask
app = Flask(__name__)
app.secret_key = 'clé_secrète_ici'
#app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://postgres:Wenjus2001%3F@localhost/soyerpp'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite://beninsoyaerp.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER_SEAL'] = 'static/images/seal'
app.config['UPLOAD_FOLDER_WEIGHT'] = 'static/images/weight'
app.config['UPLOAD_FOLDER_WEIGHT_LOCALESALE'] = 'static/images/weight_localSales'
app.config['ALLOWED_EXTENSIONS'] = {'jpg', 'jpeg', 'png', 'gif'}
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
# Initialisation de la base de données
db.init_app(app)




@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


def session_protected(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        
        user = db.session.get(User, current_user.id)
        if 'session_token' not in session or session['session_token'] != user.active_session_token:
            logout_user()
            flash("Session expired or logged in from another device.")
            return redirect(url_for('login'))
        
        return f(*args, **kwargs)
    return decorated_function


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

 






def log_action(user_id, username, action, entry_code=None):
    if entry_code is None:
        entry_code = "No entry code"
    time = datetime.now()
    new_action = LogAction(user_id=user_id, username=username, action=action, entry_code=entry_code, time=time)
    db.session.add(new_action)
    db.session.commit()


# Fonction pour récupérer les informations de l'utilisateur actuellement connecté
def get_user_info():
    if current_user.is_authenticated:
        return {
            'username': current_user.username,
            'full_name': current_user.full_name,
            'email': current_user.email,
            'location': current_user.location,
        }
    else:
        return None

# Ajoutez le contexte global pour les informations de l'utilisateur
@app.context_processor
def inject_user_info():
    return dict(user_info=get_user_info())

def role_required(*role_names):
    def decorator(func):
        @wraps(func)
        def decorated_function(*args, **kwargs):
            # Check if user is logged in and has the required role
            if current_user.is_authenticated:
                user_roles = [role.name for role in current_user.roles]
                if any(role in user_roles for role in role_names):
                    return func(*args, **kwargs)
            flash('Access not permitted! Contact the system admin for assistance.', 'error')
            return redirect(request.referrer or url_for('home'))
        return decorated_function
    return decorator 

@app.route('/')
@login_required
def home():
    """
    Affiche la page d'accueil avec les rôles de l'utilisateur et certaines données système.
    """
    roles = [role.name for role in current_user.roles] if current_user else []

    # Données de base
    total_clients = db.session.query(func.count(Customer.id)).scalar() or 0
    total_contracts = db.session.query(func.count(Contract.id)).scalar() or 0
    total_bookings = db.session.query(func.count(Booking.id)).scalar() or 0
    total_net_weight = db.session.query(func.sum(Weight.net_weight)).scalar() or 0
    total_freight_forwarders = db.session.query(func.count(FreightForwarder.id)).scalar() or 0

    # Produit le plus souvent associé à un client
    most_common_product = db.session.query(Customer.product, func.count(Customer.product)).group_by(Customer.product).order_by(func.count(Customer.product).desc()).first() or ("No Data", 0)

    # Répartition des clients par localisation
    location_distribution = db.session.query(Customer.address, func.count(Customer.id)).group_by(Customer.address).all() or []

    # Nombre total de contrats par client
    total_contracts_per_client = db.session.query(Customer.name, func.count(Contract.id)).join(Contract, Customer.id == Contract.customer_id).group_by(Customer.id).all() or []

    # Clients avec le plus grand nombre de réservations
    clients_with_most_bookings = db.session.query(Customer.name, func.count(Booking.id)).join(Contract, Customer.id == Contract.customer_id).join(Booking, Contract.id == Booking.contract_id).group_by(Customer.id).order_by(func.count(Booking.id).desc()).all() or []

    # Répartition des contrats par usine
    contracts_per_plant = db.session.query(Contract.plant, func.count(Contract.id)).group_by(Contract.plant).all() or []

    # Pourcentage moyen de la quantité contractuelle chargée
    average_percentage_loaded = db.session.query(func.avg((Contract.quantity_loaded / Contract.booking_planned) * 100)).filter(Contract.quantity_loaded.isnot(None)).scalar() or 0

    # Contrats avec le plus grand nombre de réservations en cours
    contracts_with_most_bookings = db.session.query(Contract.id, func.count(Booking.id)).join(Booking, Contract.id == Booking.contract_id).filter(Booking.container_loaded < Booking.container_planned).group_by(Contract.id).order_by(func.count(Booking.id).desc()).all() or []

    # Pourcentage de réservations chargées
    percentage_loaded_bookings = db.session.query(func.avg((Booking.container_loaded / Booking.container_planned) * 100)).filter(Booking.container_loaded.isnot(None)).scalar() or 0

    # Produit le plus fréquemment réservé
    most_common_reserved_product = db.session.query(Booking.product, func.count(Booking.product)).group_by(Booking.product).order_by(func.count(Booking.product).desc()).first() or ("No Data", 0)

    # Nombre de conteneurs associés à une réservation donnée
    containers_per_booking = db.session.query(Booking.booking_name, func.count(Container.id)).join(Container, Booking.id == Container.booking_id).group_by(Booking.booking_name).all() or []

    # Répartition des conteneurs par statut
    containers_status_distribution = db.session.query(
        case(
            (Container.loading == None, 'In Loading'),
            (Container.seal == None, 'Waiting for Seal'),
            (Container.weight == None, 'Waiting for Weight'),
            else_='Completed'
        ).label('status'),
        func.count(Container.id)
    ).group_by(
        case(
            (Container.loading == None, 'In Loading'),
            (Container.seal == None, 'Waiting for Seal'),
            (Container.weight == None, 'Waiting for Weight'),
            else_='Completed'
        )
    ).all() or []

    # Transitaire avec le plus grand nombre de réservations
    forwarders_with_most_bookings = db.session.query(FreightForwarder.name, func.count(Booking.id)).join(Booking, FreightForwarder.id == Booking.freight_forwarder_id).group_by(FreightForwarder.id).order_by(func.count(Booking.id).desc()).all() or []

    # Création du tableau croisé dynamique
    query = db.session.query(
        Customer.name.label('customer_name'),
        Contract.contract_number.label('contract_number'),
        Booking.booking_name.label('booking_name'),
        Container.container_name.label('container_name'),
        Weight.net_weight.label('net_weight')
    ).join(Contract, Customer.id == Contract.customer_id).join(Booking, Contract.id == Booking.contract_id).join(Container, Booking.id == Container.booking_id).join(Weight, Container.id == Weight.container_id).all()

    df = pd.DataFrame(query, columns=['customer_name', 'contract_number', 'booking_name', 'container_name', 'net_weight'])
    pivot_table = pd.pivot_table(df, values='net_weight', index=['customer_name', 'contract_number', 'booking_name'], columns='container_name', aggfunc='sum', fill_value=0)
    pivot_table_html = pivot_table.to_html()

    return render_template(
        'index.html',
        username=current_user.username,
        roles=roles,
        total_clients=total_clients,
        most_common_product=most_common_product,
        location_distribution=location_distribution,
        total_contracts_per_client=total_contracts_per_client,
        clients_with_most_bookings=clients_with_most_bookings,
        total_contracts=total_contracts,
        contracts_per_plant=contracts_per_plant,
        average_percentage_loaded=average_percentage_loaded,
        contracts_with_most_bookings=contracts_with_most_bookings,
        total_bookings=total_bookings,
        percentage_loaded_bookings=percentage_loaded_bookings,
        most_common_reserved_product=most_common_reserved_product,
        containers_per_booking=containers_per_booking,
        containers_status_distribution=containers_status_distribution,
        total_net_weight=total_net_weight,
        total_freight_forwarders=total_freight_forwarders,
        forwarders_with_most_bookings=forwarders_with_most_bookings,
        pivot_table_html=pivot_table_html
    )

@app.route('/export-logs-to-excel')
@login_required
@role_required('System Administrator')
def export_logs_to_excel():
    """
    Exporte les journaux système au format Excel.
    Accessible uniquement aux utilisateurs avec le rôle 'System Administrator'.
    """
    logs = LogAction.query.all()
    log_data = {
        'User ID': [log.user_id for log in logs],
        'Username': [log.username for log in logs],
        'Action': [log.action for log in logs],
        'Entry Code': [log.entry_code for log in logs],
        'Timestamp': [log.time.strftime("%Y-%m-%d %H:%M:%S") for log in logs],
    }
    df = pd.DataFrame(log_data)
    excel_file = BytesIO()
    df.to_excel(excel_file, index=False)
    excel_file.seek(0)
    response = make_response(excel_file.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=logs.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response




#--------------------Authentification-------------------------------------------------#





@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password, password):
            if user.active_session_token:
                log_action(user.id, user.username, 'logged out from other device')
                logout_user_from_other_device(user)
            token = secrets.token_urlsafe()
            user.active_session_token = token
            db.session.commit()
            login_user(user)
            session['session_token'] = token
            log_action(current_user.id, current_user.username, 'logged in')
            flash('Logged in successfully.', 'success')
            return redirect(url_for('home'))
        else:
            flash('Invalid username or password.', 'error')
    return render_template('admin/authentification/login.html')

def logout_user_from_other_device(user):
    user.active_session_token = None
    db.session.commit()

@app.route('/logout', methods=['GET', 'POST'])
@login_required
def logout():
    user = current_user
    if user.is_authenticated:
        user.active_session_token = None
        db.session.commit()
        log_action(user.id, user.username, 'logged out')
        logout_user()
        flash('Logged out successfully.', 'success')
    return redirect(url_for('home'))

@app.teardown_request
def teardown_request(exception=None):
    if current_user.is_authenticated and 'session_token' in session:
        user = db.session.get(User, current_user.id)
        if user and user.active_session_token != session['session_token']:
            logout_user()
            session.pop('session_token', None)
            flash('You have been logged out due to login from another device.', 'warning')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        full_name = request.form['full_name']
        location = request.form['location']
        position = request.form['position']
        username = request.form['username']
        email = request.form['email']
        password = request.form['password']

        # Validate input
        if not validate_full_name(full_name):
            flash('Invalid full name format. Please enter first name and last name.', 'error')
            return redirect(url_for('register'))

        if not validate_email(email):
            flash('Invalid email format.', 'error')
            return redirect(url_for('register'))

        if not validate_password(password):
            flash('Password must be at least 6 characters long and include one uppercase letter, one lowercase letter, one digit, and one special character.', 'error')
            return redirect(url_for('register'))

        # Check if username or email already exists
        existing_user = User.query.filter((User.username == username) | (User.email == email)).first()

        if existing_user:
            if existing_user.username == username:
                flash('Username already exists.', 'error')
            if existing_user.email == email:
                flash('Email already in use.', 'error')
            return redirect(url_for('register'))

        # Create new user
        hashed_password = generate_password_hash(password)

        if User.query.count() == 0:
            admin_role = Role(name='System Administrator')
            db.session.add(admin_role)
            db.session.commit()  # Commit to get the role ID for new_user
            new_user = User(full_name=full_name, location=location, position=position, username=username, email=email, password=hashed_password)
            new_user.roles.append(admin_role)
        else:
            new_user = User(username=username, email=email, password=hashed_password, location=location, position=position, full_name=full_name)

        db.session.add(new_user)
        
        try:
            db.session.commit()
            flash('Registration successful!', 'success')
            return redirect(url_for('login'))
        except IntegrityError:
            db.session.rollback()
            flash('An error occurred while processing your registration. Please try again.', 'error')
            return redirect(url_for('register'))
    
    return render_template('admin/authentification/register.html')

@app.route('/change_password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        username = request.form['username']
        old_password = request.form['old_password']
        new_password = request.form['new_password']
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password, old_password):
            if not validate_password(new_password):
                flash('Password must be at least 6 characters long and include one uppercase letter, one lowercase letter, one digit, and one special character.', 'error')
                return redirect(url_for('change_password'))
            user.password = generate_password_hash(new_password)
            db.session.commit()
            flash('Password changed successfully!', 'success')
            return redirect(url_for('login'))
        else:
            flash('Incorrect username or old password!', 'error')
    return render_template('admin/authentification/change_password.html')

# Ajoutez ici d'autres routes et fonctionnalités, en utilisant @login_required selon les besoins.


#-----------------------------Crud User for System------------------------------------------------



@app.route('/users')
@login_required
@role_required('System Administrator')
def list_users():
    """
    Affiche la liste de tous les utilisateurs.
    Accessible uniquement aux utilisateurs avec le rôle 'System Administrator'.
    """
    users = User.query.all()
    log_action(current_user.id, current_user.username, 'list users')
    return render_template('admin/user/list_users.html', users=users)

@app.route('/users/add', methods=['GET', 'POST'])
@login_required
@role_required('System Administrator')
def add_user():
    """
    Permet l'ajout d'un nouvel utilisateur.
    Accessible uniquement aux utilisateurs authentifiés avec le rôle 'System Administrator'.
    """
    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        password = request.form['password']
        location = request.form['location']
        position = request.form['position']
        full_name = request.form['full_name']

        if not validate_email(email):
            flash('Adresse email invalide !', 'error')
            return redirect(url_for('add_user'))

        if not validate_password(password):
            flash('Le mot de passe doit contenir au moins 6 caractères, une lettre majuscule, une lettre minuscule, un chiffre et un caractère spécial.', 'error')
            return redirect(url_for('add_user'))

        hashed_password = generate_password_hash(password)
        new_user = User(username=username, position=position, email=email, password=hashed_password, location=location, full_name=full_name)
        db.session.add(new_user)
        db.session.commit()
        log_action(current_user.id, current_user.username, 'User added')
        flash('Utilisateur ajouté avec succès !', 'success')
        return redirect(url_for('list_users'))
    
    return render_template('admin/user/add_user.html')

@app.route('/users/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@role_required('System Administrator')
def edit_user(id):
    """
    Permet la modification des informations d'un utilisateur spécifique.
    Accessible uniquement aux utilisateurs avec le rôle 'System Administrator'.
    """
    user = User.query.get_or_404(id)
    
    if request.method == 'POST':
        new_username = request.form['username']
        new_email = request.form['email']
        new_password = request.form['password']
        new_location = request.form['location']
        new_position = request.form['position']
        new_full_name = request.form['full_name']

        if not validate_email(new_email):
            flash('Adresse email invalide !', 'error')
            return redirect(url_for('edit_user', id=id))

        if new_password and not validate_password(new_password):
            flash('Le mot de passe doit contenir au moins 6 caractères, une lettre majuscule, une lettre minuscule, un chiffre et un caractère spécial.', 'error')
            return redirect(url_for('edit_user', id=id))

        user.username = new_username
        user.email = new_email
        user.location = new_location
        user.position = new_position
        user.full_name = new_full_name

        if new_password:
            user.password = generate_password_hash(new_password)

        db.session.commit()
        log_action(current_user.id, current_user.username, 'User Edited')
        flash('Utilisateur mis à jour avec succès !', 'success')
        return redirect(url_for('list_users'))
    
    return render_template('admin/user/edit_user.html', user=user)

@app.route('/users/delete/<int:id>', methods=['POST'])
@login_required
@role_required('System Administrator')
def delete_user(id):
    """
    Permet la suppression d'un utilisateur spécifique.
    Accessible uniquement aux utilisateurs avec le rôle 'System Administrator'.
    """
    user = User.query.get_or_404(id)
    db.session.delete(user)
    db.session.commit()
    log_action(current_user.id, current_user.username, 'User deleted')
    flash('Utilisateur supprimé avec succès !', 'success')
    return redirect(url_for('list_users'))

#-------------------------------CRUD Roles for Systems --------------------------------------------

@app.route('/roles')
@login_required
@role_required('System Administrator')
def list_roles():
    """
    Affiche la liste de tous les rôles disponibles.
    Accessible uniquement aux utilisateurs authentifiés avec les rôles 'System Administrator' ou 'Logistics Manager'.
    """
    roles = Role.query.all()
    log_action(current_user.id, current_user.username, 'Roles list')
    return render_template('admin/roles/list_roles.html', roles=roles)

@app.route('/roles/add', methods=['GET', 'POST'])
@login_required
@role_required('System Administrator')
def add_role():
    """
    Permet l'ajout d'un nouveau rôle.
    Accessible uniquement aux utilisateurs authentifiés avec le rôle 'System Administrator'.
    """
    if request.method == 'POST':
        name = request.form['name']
        new_role = Role(name=name)
        db.session.add(new_role)
        db.session.commit()
        log_action(current_user.id, current_user.username, 'Role added')
        flash('Role added successfully!', 'success')
        return redirect(url_for('list_roles'))
    
    return render_template('admin/roles/add_role.html')

@app.route('/roles/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@role_required('System Administrator')
def edit_role(id):
    """
    Permet la modification d'un rôle spécifique.
    Accessible uniquement aux utilisateurs authentifiés avec le rôle 'System Administrator'.
    """
    role = Role.query.get_or_404(id)
    
    if request.method == 'POST':
        role.name = request.form['name']
        db.session.commit()
        log_action(current_user.id, current_user.username, 'Role edited')
        flash('Role updated successfully!', 'success')
        return redirect(url_for('list_roles'))
    
    return render_template('admin/roles/edit_role.html', role=role)

@app.route('/roles/delete/<int:id>', methods=['POST'])
@login_required
@role_required('System Administrator')
def delete_role(id):
    """
    Permet la suppression d'un rôle spécifique.
    Accessible uniquement aux utilisateurs authentifiés avec le rôle 'System Administrator'.
    """
    role = Role.query.get_or_404(id)
    db.session.delete(role)
    db.session.commit()
    log_action(current_user.id, current_user.username, 'Role deleted')
    flash('Role deleted successfully!', 'success')
    return redirect(url_for('list_roles'))

@app.route('/users/assign_role/<int:user_id>', methods=['GET', 'POST'])
@login_required
@role_required('System Administrator')
def assign_role(user_id):
    """
    Permet l'attribution de rôles à un utilisateur spécifique.
    Accessible uniquement aux utilisateurs authentifiés avec les rôles 'System Administrator' ou 'Logistics Manager'.
    """
    user = User.query.get_or_404(user_id)
    roles = Role.query.all()
    
    if request.method == 'POST':
        selected_roles = request.form.getlist('roles')
        user.roles = Role.query.filter(Role.id.in_(selected_roles)).all()
        db.session.commit()
        log_action(current_user.id, current_user.username, 'Role assignment')
        flash('Roles assigned successfully!', 'success')
        return redirect(url_for('list_users'))
    
    return render_template('admin/roles/assign_role.html', user=user, roles=roles)

@app.route('/users/edit_roles/<int:user_id>', methods=['GET', 'POST'])
@login_required
@role_required('System Administrator')
def edit_roles(user_id):
    """
    Permet la modification des rôles d'un utilisateur spécifique.
    Accessible uniquement aux utilisateurs authentifiés avec les rôles 'System Administrator' ou 'Logistics Manager'.
    """
    user = User.query.get_or_404(user_id)
    roles = Role.query.all()
    
    if request.method == 'POST':
        selected_roles = request.form.getlist('roles')
        user.roles = Role.query.filter(Role.id.in_(selected_roles)).all()
        db.session.commit()
        log_action(current_user.id, current_user.username, 'Role edited')
        flash('User roles updated successfully!', 'success')
        return redirect(url_for('list_users'))
    
    return render_template('admin/edit_user_roles.html', user=user, roles=roles)

@app.route('/users/remove_role/<int:user_id>', methods=['GET', 'POST'])
@login_required
@role_required('System Administrator')
def remove_role(user_id):
    """
    Permet la suppression de rôles d'un utilisateur spécifique.
    Accessible uniquement aux utilisateurs authentifiés avec le rôle 'System Administrator'.
    """
    if request.method == 'POST':
        user = User.query.get_or_404(user_id)
        selected_roles = request.form.getlist('roles')
        user.roles = [role for role in user.roles if str(role.id) not in selected_roles]
        db.session.commit()
        log_action(current_user.id, current_user.username, 'Role removed')
        flash('Role removed successfully!', 'success')
        return redirect(url_for('list_users'))
    
    user = User.query.get_or_404(user_id)
    return render_template('admin/roles/remove_role.html', user=user)

# ---------------------------------------------------HERE IS ALL OPERATION ABOUT LOGISTICS MANAGEMENT SYSTEMS---------------------------------------------#

# Variable globale pour stocker temporairement les données du client avant confirmation
temp_customer_data = None

@app.route('/add_customer', methods=['GET', 'POST'])
@login_required
@role_required('System Administrator', 'Sales Manager')
def add_customer():
    """
    Permet d'ajouter un nouveau client en fournissant les détails nécessaires.
    Accessible uniquement aux utilisateurs authentifiés avec les rôles 'System Administrator' ou 'Sales Manager'.
    """
    global temp_customer_data
    
    if request.method == 'POST':
        name = request.form.get('name')
        email = request.form.get('email')
        phone = request.form.get('phone')
        address = request.form.get('address')
        product = request.form.get('product')
        plant = request.form.get('plant')

        if not validate_email(email):
            flash('Invalid email address!', 'error')
            return redirect(url_for('add_customer'))
        
        if not validate_phone_number(phone):
            flash('Invalid phone number!', 'error')
            return redirect(url_for('add_customer'))
        
        existing_customer = Customer.query.filter((Customer.name == name) | (Customer.email == email)).first()
        if existing_customer:
            flash('A customer with this name or email already exists!', 'error')
            return redirect(url_for('add_customer'))

        temp_customer_data = {
            'name': name,
            'email': email,
            'phone': phone,
            'address': address,
            'product': product,
            'plant': plant,
            'created_by': current_user.username
        }
        return redirect(url_for('confirm_customer_creation'))
    
    return render_template('customer/add_customer.html')

@app.route('/confirm_customer_creation', methods=['GET', 'POST'])
@login_required
@role_required('System Administrator', 'Sales Manager')
def confirm_customer_creation():
    """
    Affiche une page de confirmation avant la création du client.
    Accessible uniquement aux utilisateurs authentifiés avec les rôles 'System Administrator' ou 'Sales Manager'.
    """
    global temp_customer_data

    if request.method == 'POST':
        confirm_choice = request.form.get('choice')

        if confirm_choice == 'confirm':
            new_customer = Customer(
                name=temp_customer_data['name'],
                email=temp_customer_data['email'],
                phone=temp_customer_data['phone'],
                address=temp_customer_data['address'],
                product=temp_customer_data['product'],
                plant=temp_customer_data['plant'],
                created_by=current_user.username
            )

            db.session.add(new_customer)
            db.session.commit()

            flash('Customer added successfully', 'success')
        elif confirm_choice == 'cancel':
            flash('Customer addition canceled.', 'info')

        temp_customer_data = None
        return redirect(url_for('customers'))

    if temp_customer_data:
        customer_data = temp_customer_data
    else:
        return redirect(url_for('add_customer'))

    return render_template('customer/confirm_customer_creation.html', customer_data=customer_data)

@app.route('/customer/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('System Administrator', 'Sales Manager')
def edit_customer(id):
    """
    Permet de modifier les informations d'un client spécifique.
    Accessible uniquement aux utilisateurs authentifiés avec les rôles 'System Administrator' ou 'Sales Manager'.
    """
    customer = Customer.query.get_or_404(id)
    
    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        phone = request.form['phone']
        address = request.form['address']
        product = request.form['product']
        plant = request.form['plant']

        if not validate_email(email):
            flash('Invalid email address!', 'error')
            return redirect(url_for('edit_customer', id=id))
        
        if not validate_phone_number(phone):
            flash('Invalid phone number!', 'error')
            return redirect(url_for('edit_customer', id=id))
        
        customer.name = name
        customer.email = email
        customer.phone = phone
        customer.address = address
        customer.product = product
        customer.plant = plant
        customer.created_by = current_user.username

        db.session.commit()
        flash('Customer updated successfully!', 'success')
        
        return redirect(url_for('customers'))
    
    return render_template('customer/edit_customer.html', customer=customer)

@app.route('/customer/<int:id>/delete', methods=['POST'])
@login_required
@role_required('System Administrator')
def delete_customer(id):
    """
    Permet de supprimer un client spécifique.
    Accessible uniquement aux utilisateurs authentifiés avec le rôle 'System Administrator'.
    """
    customer = Customer.query.get_or_404(id)
    
    contracts = Contract.query.filter_by(customer_id=id).all()
    if contracts:
        flash(f'Cannot delete customer {customer.name} because there are associated contracts.', 'error')
    else:
        db.session.delete(customer)
        db.session.commit()
        flash(f'Customer {customer.name} deleted successfully', 'warning')
    
    return redirect(url_for('customers'))

@app.route('/customers')
@login_required
@role_required('System Administrator', 'Sales Manager', 'Customer Support')
def customers():
    """
    Affiche la liste de tous les clients.
    Accessible uniquement aux utilisateurs authentifiés avec les rôles 'System Administrator', 'Sales Manager' ou 'Customer Support'.
    """
    customers = Customer.query.all()
    
    customers_list = [
        {
            'id': customer.id,
            'name': customer.name,
            'email': customer.email,
            'plant': customer.plant,
            'product': customer.product
        }
        for customer in customers
    ]
    
    return render_template('customer/customers.html', customers=customers_list)

@app.route('/customer/<int:customer_id>/view', methods=['GET'])
@login_required
@role_required('System Administrator', 'Sales Manager', 'Customer Support')
def view_customer(customer_id):
    """
    Affiche les détails d'un client spécifique.
    Accessible uniquement aux utilisateurs authentifiés avec les rôles 'System Administrator', 'Sales Manager' ou 'Customer Support'.
    """
    customer = Customer.query.get_or_404(customer_id)
    
    return render_template('customer/view_customer.html', customer=customer)



#----------------------Freight Forwarders-------------------------------------------------------------------
# Variable globale pour stocker temporairement les données du Freight Forwarder avant confirmation

# Variable globale pour stocker temporairement les données du Freight Forwarder avant confirmation
temp_freight_forwarder_data = None

@app.route('/add_freight_forwarder', methods=['GET', 'POST'])
@login_required
@role_required('System Administrator', 'Logistics Manager', 'Transportation Coordinator')
def add_freight_forwarder():
    global temp_freight_forwarder_data
    
    if request.method == 'POST':
        name = request.form.get('name')
        tel = request.form.get('tel')

        if not name or not name.isalpha():
            flash('Invalid name! Only alphabetic characters are allowed.', 'error')
            return redirect(url_for('add_freight_forwarder'))
        
        if not validate_phone_number(tel):
            flash('Invalid phone number!', 'error')
            return redirect(url_for('add_freight_forwarder'))

        existing_forwarder = FreightForwarder.query.filter_by(name=name).first()
        if existing_forwarder:
            flash('Freight Forwarder already exists!', 'error')
            return redirect(url_for('add_freight_forwarder'))

        temp_freight_forwarder_data = {
            'name': name,
            'tel': tel,
            'freight_created_by': current_user.username
        }
        return redirect(url_for('confirm_freight_forwarder_creation'))
    
    return render_template('freight_forwarder/add_freight_forwarder.html')

@app.route('/confirm_freight_forwarder_creation', methods=['GET', 'POST'])
@login_required
@role_required('System Administrator', 'Logistics Manager', 'Transportation Coordinator')
def confirm_freight_forwarder_creation():
    global temp_freight_forwarder_data

    if request.method == 'POST':
        confirm_choice = request.form.get('choice')

        if confirm_choice == 'confirm':
            new_freight_forwarder = FreightForwarder(
                name=temp_freight_forwarder_data['name'],
                tel=temp_freight_forwarder_data['tel'],
                freight_created_by=current_user.username
            )

            db.session.add(new_freight_forwarder)
            db.session.commit()

            flash('Freight Forwarder added successfully', 'success')
        elif confirm_choice == 'cancel':
            flash('Freight Forwarder addition canceled.', 'info')

        temp_freight_forwarder_data = None
        return redirect(url_for('freight_forwarders'))

    if temp_freight_forwarder_data:
        freight_forwarder_data = temp_freight_forwarder_data
    else:
        return redirect(url_for('add_freight_forwarder'))

    return render_template('freight_forwarder/confirm_freight_forwarder_creation.html', freight_forwarder_data=freight_forwarder_data)

@app.route('/freight_forwarder/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('System Administrator', 'Logistics Manager', 'Transportation Coordinator')
def edit_freight_forwarder(id):
    freight_forwarder = FreightForwarder.query.get_or_404(id)

    if request.method == 'POST':
        name = request.form['name']
        tel = request.form['tel']

        if not name or not name.isalpha():
            flash('Invalid name! Only alphabetic characters are allowed.', 'error')
            return redirect(url_for('edit_freight_forwarder', id=id))
        
        if not validate_phone_number(tel):
            flash('Invalid phone number!', 'error')
            return redirect(url_for('edit_freight_forwarder', id=id))

        freight_forwarder.name = name
        freight_forwarder.tel = tel
        freight_forwarder.freight_created_by = current_user.username

        db.session.commit()
        flash('Freight Forwarder updated successfully', 'success')
        
        return redirect(url_for('freight_forwarders'))
    
    return render_template('freight_forwarder/edit_freight_forwarder.html', freight_forwarder=freight_forwarder)

@app.route('/freight_forwarder/<int:id>/delete', methods=['POST'])
@login_required
@role_required('System Administrator')
def delete_freight_forwarder(id):
    freight_forwarder = FreightForwarder.query.get_or_404(id)
    
    bookings = Booking.query.filter_by(freight_forwarder_id=id).all()
    
    if bookings:
        flash(f'Cannot delete Freight Forwarder {freight_forwarder.name} because there are associated bookings.', 'error')
    else:
        db.session.delete(freight_forwarder)
        db.session.commit()
        flash(f'Freight Forwarder {freight_forwarder.name} deleted successfully', 'warning')
    
    return redirect(url_for('freight_forwarders'))

@app.route('/freight_forwarders')
@login_required
@role_required('System Administrator', 'Logistics Manager', 'Customer Support Representative', 'Logistics Data Analyst')
def freight_forwarders():
    freight_forwarders = FreightForwarder.query.all()
    
    freight_forwarders_list = [
        {
            'id': freight_forwarder.id,
            'name': freight_forwarder.name,
            'tel': freight_forwarder.tel,
            'freight_created_at': freight_forwarder.freight_created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'freight_created_by': freight_forwarder.freight_created_by
        }
        for freight_forwarder in freight_forwarders
    ]
    
    return render_template('freight_forwarder/freight_forwarders.html', freight_forwarders=freight_forwarders_list)

@app.route('/freight_forwarder/<int:id>/view')
@login_required
@role_required('System Administrator', 'Logistics Manager', 'Customer Support Representative')
def view_freight_forwarder(id):
    freight_forwarder = FreightForwarder.query.get_or_404(id)
    
    return render_template('freight_forwarder/view_freight_forwarder.html', freight_forwarder=freight_forwarder)



#---------------Contract Management -----------------------------------------------------------------------

# Variable globale pour stocker temporairement les données du contrat avant confirmation
# Variable globale pour stocker temporairement les données du contrat avant confirmation
temp_contract_data = None

@app.route('/add_contract', methods=['GET', 'POST'])
@login_required
@role_required('System Administrator', 'Logistics Manager', 'Purchasing Manager', 'Manager')
def add_contract():
    """
    Permet d'ajouter un nouveau contrat en fournissant les détails nécessaires.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    global temp_contract_data

    if request.method == 'POST':
        contract_number = request.form.get('contract_number')
        plant = request.form.get('plant')
        product = request.form.get('product')
        customer_id = request.form.get('customer_id')
        shipment_start_date = request.form.get('shipment_start_date')
        shipment_end_date = request.form.get('shipment_end_date')
        destination = request.form.get('destination')
        contract_qty = request.form.get('contract_qty')
        bag_type = request.form.get('bag_type')
        booking_planned = request.form.get('booking_planned')

        if not contract_number or not validate_contract_number(contract_number):
            flash('Invalid contract number!', 'error')
            return redirect(url_for('add_contract'))

        existing_contract = Contract.query.filter_by(contract_number=contract_number).first()
        if existing_contract:
            flash('Contract number already exists!', 'error')
            return redirect(url_for('add_contract'))

        try:
            start_date = datetime.strptime(shipment_start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(shipment_end_date, '%Y-%m-%d').date()
            if start_date > end_date:
                flash('The shipment start date cannot be after the end date!', 'error')
                return redirect(url_for('add_contract'))
        except ValueError:
            flash('Invalid date format! Please use YYYY-MM-DD.', 'error')
            return redirect(url_for('add_contract'))

        if not validate_number(contract_qty):
            flash('Invalid contract quantity!', 'error')
            return redirect(url_for('add_contract'))

        temp_contract_data = {
            'contract_number': contract_number,
            'plant': plant,
            'product': product,
            'customer_id': customer_id,
            'shipment_start_date': shipment_start_date,
            'shipment_end_date': shipment_end_date,
            'destination': destination,
            'contract_qty': contract_qty,
            'bag_type': bag_type,
            'booking_planned': booking_planned,
            'created_by': current_user.username
        }

        return redirect(url_for('confirm_contract_creation'))

    customers = Customer.query.all()
    return render_template('contract/add_contract.html', customers=customers)

@app.route('/confirm_contract_creation', methods=['GET', 'POST'])
@login_required
@role_required('System Administrator', 'Logistics Manager', 'Purchasing Manager', 'Manager')
def confirm_contract_creation():
    """
    Affiche une page de confirmation avant la création du contrat.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    global temp_contract_data

    if request.method == 'POST':
        confirm_choice = request.form.get('choice')

        if confirm_choice == 'confirm':
            new_contract = Contract(
                contract_number=temp_contract_data['contract_number'],
                plant=temp_contract_data['plant'],
                product=temp_contract_data['product'],
                customer_id=temp_contract_data['customer_id'],
                shipment_start_date=datetime.strptime(temp_contract_data['shipment_start_date'], '%Y-%m-%d').date(),
                shipment_end_date=datetime.strptime(temp_contract_data['shipment_end_date'], '%Y-%m-%d').date(),
                destination=temp_contract_data['destination'],
                contract_qty=temp_contract_data['contract_qty'],
                bag_type=temp_contract_data['bag_type'],
                booking_planned=temp_contract_data['booking_planned'],
                created_by=current_user.username
            )

            db.session.add(new_contract)
            db.session.commit()

            flash('Contract added successfully', 'success')
        elif confirm_choice == 'cancel':
            flash('Contract addition canceled.', 'info')

        temp_contract_data = None

        return redirect(url_for('contracts'))

    if temp_contract_data:
        contract_data = temp_contract_data
        customer = Customer.query.get(contract_data['customer_id'])
        contract_data['customer_name'] = customer.name if customer else 'Unknown'
    else:
        return redirect(url_for('add_contract'))

    return render_template('contract/confirm_contract_creation.html', contract_data=contract_data)

@app.route('/contract/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('System Administrator', 'Logistics Manager', 'Purchasing Manager', 'Manager')
def edit_contract(id):
    """
    Permet de modifier les informations d'un contrat spécifique.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    contract = Contract.query.get_or_404(id)

    if request.method == 'POST':
        contract_number = request.form['contract_number']
        plant = request.form['plant']
        product = request.form['product']
        customer_id = request.form['customer_id']
        shipment_start_date = request.form['shipment_start_date']
        shipment_end_date = request.form['shipment_end_date']
        destination = request.form['destination']
        contract_qty = request.form['contract_qty']
        bag_type = request.form['bag_type']

        if not contract_number or not validate_contract_number(contract_number):
            flash('Invalid contract number!', 'error')
            return redirect(url_for('edit_contract', id=id))

        try:
            start_date = datetime.strptime(shipment_start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(shipment_end_date, '%Y-%m-%d').date()
            if start_date > end_date:
                flash('The shipment start date cannot be after the end date!', 'error')
                return redirect(url_for('edit_contract', id=id))
        except ValueError:
            flash('Invalid date format! Please use YYYY-MM-DD.', 'error')
            return redirect(url_for('edit_contract', id=id))

        if not validate_number(contract_qty):
            flash('Invalid contract quantity!', 'error')
            return redirect(url_for('edit_contract', id=id))

        existing_contract = Contract.query.filter_by(contract_number=contract_number).first()
        if existing_contract and existing_contract.id != id:
            flash('Contract number already exists!', 'error')
            return redirect(url_for('edit_contract', id=id))

        contract.contract_number = contract_number
        contract.plant = plant
        contract.product = product
        contract.customer_id = customer_id
        contract.shipment_start_date = start_date
        contract.shipment_end_date = end_date
        contract.destination = destination
        contract.contract_qty = contract_qty
        contract.bag_type = bag_type
        contract.created_by = current_user.username

        db.session.commit()
        flash('Contract updated successfully', 'success')

        return redirect(url_for('contracts'))

    customers = Customer.query.all()
    return render_template('contract/edit_contract.html', contract=contract, customers=customers)

@app.route('/contract/<int:id>/delete', methods=['POST'])
@login_required
@role_required('System Administrator')
def delete_contract(id):
    """
    Permet de supprimer un contrat spécifique.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    contract = Contract.query.get_or_404(id)
    bookings = Booking.query.filter_by(contract_id=id).all()

    if bookings:
        flash(f'Cannot delete Contract {contract.contract_number} because there are associated bookings.', 'error')
    else:
        db.session.delete(contract)
        db.session.commit()
        flash(f'Contract {contract.contract_number} deleted successfully', 'warning')

    return redirect(url_for('contracts'))

@app.route('/contracts')
#@login_required
#@role_required('System Administrator', 'Logistics Manager', 'Purchasing Manager', 'Manager', 'Logistics Data Analyst')
def contracts():
    """
    Affiche la liste de tous les contrats.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    contracts = Contract.query.all()

    contracts_list = [
        {
            'id': contract.id,
            'plant': contract.plant,
            'product': contract.product,
            'contract_number': contract.contract_number,
            'customer_id': contract.customer_id,
            'customer_name': contract.customer.name if contract.customer else None,
            'shipment_start_date': contract.shipment_start_date.strftime('%Y-%m-%d'),
            'shipment_end_date': contract.shipment_end_date.strftime('%Y-%m-%d'),
            'destination': contract.destination,
            'contract_qty': contract.contract_qty,
            'booking_planned': contract.booking_planned,
            'created_at': contract.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'created_by': contract.created_by,
            'status': contract.status
        }
        for contract in contracts
    ]

    return render_template('contract/contracts.html', contracts=contracts_list)

@app.route('/contract/<int:id>/view')
@login_required
@role_required('System Administrator', 'Logistics Manager', 'Purchasing Manager', 'Manager', 'Customer Support Representative')
def view_contract(id):
    """
    Affiche les détails d'un contrat spécifique, y compris les bookings associés.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    contract = Contract.query.get_or_404(id)
    bookings = Booking.query.filter_by(contract_id=id).all()

    payment_info = contract.payment_info
    date_info = contract.date_info
    remarks_info = contract.remarks_info

    bookings_list = [
        {
            'id': booking.id,
            'plant': booking.plant,
            'product': booking.product,
            'contract_id': booking.contract_id,
            'booking_name': booking.booking_name,
            'bag_type': booking.bag_type,
            'container_planned': booking.container_planned,
            'quantity_planned': booking.quantity_planned,
            'booking_created_at': booking.booking_created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'booking_created_by': booking.booking_created_by,
            'freight_forwarder_id': booking.freight_forwarder_id,
            'freight_forwarder_name': booking.freight_forwarder.name if booking.freight_forwarder else 'Unknown',
            'status': booking.status
        }
        for booking in bookings
    ]

    return render_template('contract/view_contract.html', contract=contract, bookings=bookings_list, 
                           payment_info=payment_info, date_info=date_info, 
                           remarks_info=remarks_info)




#------------------------------------------------Booking Mnagement ----------------------------------------------
# Variable globale pour stocker temporairement les données de réservation avant confirmation
# Variable globale pour stocker temporairement les données de réservation avant confirmation
temp_booking_data = None

@app.route('/booking/add/<int:contract_id>', methods=['GET', 'POST'])
@login_required
@role_required('Logistics Manager', 'Purchasing Manager', 'Manager','System Administrator')
def add_booking(contract_id):
    """
    Permet d'ajouter une nouvelle réservation pour un contrat spécifique.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    global temp_booking_data

    if request.method == 'POST':
        plant = request.form['plant']
        product = request.form['product']
        booking_name = request.form['booking_name']
        bag_type = request.form['bag_type']
        container_planned = request.form['container_planned']
        quantity_planned = request.form['quantity_planned']
        freight_forwarder_id = request.form['freight_forwarder_id']

        if not validate_number(quantity_planned):
            flash('Invalid quantity planned!', 'error')
            return redirect(url_for('add_booking', contract_id=contract_id))

        if not Contract.query.get(contract_id):
            flash('Invalid contract ID!', 'error')
            return redirect(url_for('add_booking', contract_id=contract_id))

        if not FreightForwarder.query.get(freight_forwarder_id):
            flash('Invalid freight forwarder ID!', 'error')
            return redirect(url_for('add_booking', contract_id=contract_id))

        existing_booking = Booking.query.filter_by(booking_name=booking_name).first()
        if existing_booking:
            flash('Booking name already exists!', 'error')
            return redirect(url_for('add_booking', contract_id=contract_id))

        temp_booking_data = {
            'plant': plant,
            'product': product,
            'contract_id': contract_id,
            'booking_name': booking_name,
            'bag_type': bag_type,
            'container_planned': container_planned,
            'quantity_planned': quantity_planned,
            'booking_created_by': current_user.username,
            'freight_forwarder_id': freight_forwarder_id
        }
        return redirect(url_for('confirm_booking_creation'))

    contract = Contract.query.get_or_404(contract_id)
    freight_forwarders = FreightForwarder.query.all()
    return render_template('booking/add_booking.html', contract_id=contract_id, contract=contract, freight_forwarders=freight_forwarders)

@app.route('/confirm_booking_creation', methods=['GET', 'POST'])
@login_required
@role_required('Logistics Manager', 'Purchasing Manager', 'Manager','System Administrator')
def confirm_booking_creation():
    """
    Affiche une page de confirmation avant la création de la réservation.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    global temp_booking_data

    if request.method == 'POST':
        confirm_choice = request.form.get('choice')

        if confirm_choice == 'confirm':
            new_booking = Booking(
                plant=temp_booking_data['plant'],
                product=temp_booking_data['product'],
                contract_id=temp_booking_data['contract_id'],
                booking_name=temp_booking_data['booking_name'],
                bag_type=temp_booking_data['bag_type'],
                container_planned=temp_booking_data['container_planned'],
                quantity_planned=temp_booking_data['quantity_planned'],
                booking_created_by=current_user.username,
                freight_forwarder_id=temp_booking_data['freight_forwarder_id']
            )

            db.session.add(new_booking)
            db.session.commit()

            flash('Booking added successfully', 'success')
        elif confirm_choice == 'cancel':
            flash('Booking addition canceled.', 'info')

        temp_booking_data = None
        return redirect(url_for('bookings'))

    if temp_booking_data:
        booking_data = temp_booking_data
        contract = Contract.query.get(booking_data['contract_id'])
        freight_forwarder = FreightForwarder.query.get(booking_data['freight_forwarder_id'])
        booking_data['contract_number'] = contract.contract_number if contract else 'Unknown'
        booking_data['freight_forwarder_name'] = freight_forwarder.name if freight_forwarder else 'Unknown'
    else:
        return redirect(url_for('add_booking', contract_id=temp_booking_data['contract_id']))

    return render_template('booking/confirm_booking_creation.html', booking_data=booking_data)

@app.route('/booking/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('Logistics Manager', 'Purchasing Manager', 'Manager','System Administrator')
def edit_booking(id):
    """
    Permet de modifier les informations d'une réservation spécifique.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    booking = Booking.query.get_or_404(id)

    if request.method == 'POST':
        plant = request.form['plant']
        product = request.form['product']
        contract_id = request.form['contract_id']
        booking_name = request.form['booking_name']
        bag_type = request.form['bag_type']
        container_planned = request.form['container_planned']
        quantity_planned = request.form['quantity_planned']
        freight_forwarder_id = request.form['freight_forwarder_id']

        if not validate_number(quantity_planned):
            flash('Invalid quantity planned!', 'error')
            return redirect(url_for('edit_booking', id=id))

        if not Contract.query.get(contract_id):
            flash('Invalid contract ID!', 'error')
            return redirect(url_for('edit_booking', id=id))

        if not FreightForwarder.query.get(freight_forwarder_id):
            flash('Invalid freight forwarder ID!', 'error')
            return redirect(url_for('edit_booking', id=id))

        existing_booking = Booking.query.filter_by(booking_name=booking_name).first()
        if existing_booking and existing_booking.id != id:
            flash('Booking name already exists!', 'error')
            return redirect(url_for('edit_booking', id=id))

        booking.plant = plant
        booking.product = product
        booking.contract_id = contract_id
        booking.booking_name = booking_name
        booking.bag_type = bag_type
        booking.container_planned = container_planned
        booking.quantity_planned = quantity_planned
        booking.booking_created_by = current_user.username
        booking.freight_forwarder_id = freight_forwarder_id

        db.session.commit()
        flash('Booking updated successfully', 'success')
        return redirect(url_for('bookings'))

    contracts = Contract.query.all()
    freight_forwarders = FreightForwarder.query.all()
    return render_template('booking/edit_booking.html', booking=booking, contracts=contracts, freight_forwarders=freight_forwarders)

@app.route('/booking/<int:id>/delete', methods=['POST'])
@login_required
@role_required('System Administrator')
def delete_booking(id):
    """
    Permet de supprimer une réservation spécifique.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    booking = Booking.query.get_or_404(id)
    containers = Container.query.filter_by(booking_id=id).all()

    if containers:
        flash(f'Cannot delete Booking {booking.booking_name} because there are associated containers.', 'warning')
    else:
        db.session.delete(booking)
        db.session.commit()
        flash(f'Booking {booking.booking_name} deleted successfully', 'success')

    return redirect(url_for('bookings'))

@app.route('/bookings')
#@login_required
#@role_required('Logistics Manager', 'Purchasing Manager', 'Manager', 'Logistics Data Analyst')
def bookings():
    """
    Affiche la liste de toutes les réservations.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    bookings = Booking.query.all()

    bookings_list = [
        {
            'id': booking.id,
            'plant': booking.plant,
            'product': booking.product,
            'contract_id': booking.contract_id,
            'contract_number': booking.contract.contract_number if booking.contract else 'Unknown',
            'customer_name': booking.contract.customer.name if booking.contract and booking.contract.customer else 'Unknown',
            'booking_name': booking.booking_name,
            'bag_type': booking.bag_type,
            'container_planned': booking.container_planned,
            'quantity_planned': booking.quantity_planned,
            'booking_created_at': booking.booking_created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'booking_created_by': booking.booking_created_by,
            'freight_forwarder_id': booking.freight_forwarder_id,
            'freight_forwarder_name': booking.freight_forwarder.name if booking.freight_forwarder else 'Unknown',
            'status': booking.status
        }
        for booking in bookings
    ]

    return render_template('booking/bookings.html', bookings=bookings_list)

@app.route('/booking/<int:id>/view')
@login_required
@role_required('Logistics Manager', 'Purchasing Manager', 'Manager', 'Customer Support Representative','System Administrator')
def view_booking(id):
    """
    Affiche les détails d'une réservation spécifique, y compris les containers associés.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    booking = Booking.query.get_or_404(id)
    contract = booking.contract
    containers = Container.query.filter_by(booking_id=id).all()

    containers_list = [
        {
            'id': container.id,
            'arrival_date': container.arrival_date.strftime('%Y-%m-%d'),
            'truck_number': container.truck_number,
            'container_name': container.container_name,
            'plant': container.plant,
            'product': container.product,
            'freight_forwarder': container.freight_forwarder,
            'container_tare': container.container_tare,
            'bags_type': container.bags_type,
            'container_created_at': container.container_created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'created_by': container.container_created_by,
            'status': container.get_status()
        }
        for container in containers
    ]

    return render_template('booking/view_booking.html', booking=booking, containers=containers_list, contract=contract)




#-----------------------------------------Container data --------------------------------------


# Variable globale pour stocker temporairement les données de container avant confirmation

temp_container_data = None

@app.route('/container/add/<int:booking_id>', methods=['GET', 'POST'])
@login_required
@role_required('Logistics Manager', 'Purchasing Manager', 'Transportation Coordinator', 'Warehouse Operator', 'Manager')
def add_container(booking_id):
    """
    Permet d'ajouter un nouveau conteneur pour une réservation spécifique.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    global temp_container_data

    if request.method == 'POST':
        # Récupère les informations du formulaire
        arrival_date = request.form['arrival_date']
        truck_number = request.form['truck_number']
        container_name = request.form['container_name']
        plant = request.form['plant']
        product = request.form['product']
        freight_forwarder = request.form['freight_forwarder']
        container_tare = request.form['container_tare']
        bags_type = request.form['bags_type']

        # Validation du format du numéro de camion
        if not validate_truck_number(truck_number):
            flash('Invalid truck number format! Please use one of the following formats: A1245RB, AA1234RB, A3345RB/A4567RB, etc.', 'error')
            return redirect(url_for('add_container', booking_id=booking_id))

        # Validation du format du numéro de conteneur
        if not validate_container_number(container_name):
            flash('Invalid container name format! Please use the format: ABCD1234567.', 'error')
            return redirect(url_for('add_container', booking_id=booking_id))

        # Vérification de l'existence du numéro de conteneur
        if Container.query.filter_by(container_name=container_name).first():
            flash('Container number already exists!', 'error')
            return redirect(url_for('add_container', booking_id=booking_id))

        # Validation du format de la date d'arrivée
        try:
            arrival_date = datetime.strptime(arrival_date, '%Y-%m-%dT%H:%M')
        except ValueError:
            flash('Invalid arrival date format! Please use the format YYYY-MM-DDTHH:MM.', 'error')
            return redirect(url_for('add_container', booking_id=booking_id))

        # Stocke les données du conteneur dans une variable globale temporaire pour confirmation
        temp_container_data = {
            'arrival_date': arrival_date,
            'truck_number': truck_number,
            'container_name': container_name,
            'plant': plant,
            'product': product,
            'freight_forwarder': freight_forwarder,
            'booking_id': booking_id,
            'container_tare': container_tare,
            'bags_type': bags_type,
            'created_by': current_user.username
        }
        return redirect(url_for('confirm_container_creation'))

    # Affiche le formulaire d'ajout de conteneur avec les détails de la réservation
    booking = Booking.query.filter_by(id=booking_id).first()
    now = datetime.now(pytz.timezone('Africa/Porto-Novo'))
    return render_template('container/add_container.html', booking_id=booking_id, booking=booking, now=now)

@app.route('/confirm_container_creation', methods=['GET', 'POST'])
@login_required
@role_required('Logistics Manager', 'Purchasing Manager', 'Transportation Coordinator', 'Warehouse Operator', 'Manager')
def confirm_container_creation():
    """
    Affiche une page de confirmation avant la création du conteneur.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    global temp_container_data

    if request.method == 'POST':
        confirm_choice = request.form.get('choice')

        if confirm_choice == 'confirm':
            # Crée un nouvel objet Container avec les données temporaires
            new_container = Container(
                arrival_date=temp_container_data['arrival_date'],
                truck_number=temp_container_data['truck_number'],
                container_name=temp_container_data['container_name'],
                plant=temp_container_data['plant'],
                product=temp_container_data['product'],
                freight_forwarder=temp_container_data['freight_forwarder'],
                booking_id=temp_container_data['booking_id'],
                container_tare=temp_container_data['container_tare'],
                bags_type=temp_container_data['bags_type'],
                container_created_by=current_user.username
            )

            # Ajoute le conteneur à la base de données
            db.session.add(new_container)
            db.session.commit()

            flash('Container added successfully', 'success')
        elif confirm_choice == 'cancel':
            flash('Container addition canceled.', 'info')

        # Réinitialise les données temporaires
        temp_container_data = None

        # Redirige vers la liste des conteneurs après confirmation
        return redirect(url_for('containers'))

    # Si les données temporaires n'existent pas, redirige vers la page d'ajout
    if temp_container_data:
        container_data = temp_container_data
        # Récupère les détails de la réservation
        booking = Booking.query.get(container_data['booking_id'])
        container_data['booking_name'] = booking.booking_name if booking else 'Unknown'
    else:
        return redirect(url_for('add_container'))

    # Affiche la page de confirmation de création du conteneur avec les données temporaires
    return render_template('container/confirm_container_creation.html', container_data=container_data)

@app.route('/container/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('Logistics Manager', 'Purchasing Manager', 'Transportation Coordinator', 'Warehouse Operator', 'Manager')
def edit_container(id):
    """
    Permet de modifier les informations d'un conteneur spécifique.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    container = Container.query.get_or_404(id)

    if request.method == 'POST':
        # Récupère les nouvelles informations du formulaire
        truck_number = request.form['truck_number']
        container_name = request.form['container_name']
        plant = request.form['plant']
        product = request.form['product']
        freight_forwarder = request.form['freight_forwarder']
        booking_id = request.form['booking_id']
        container_tare = request.form['container_tare']
        bags_type = request.form['bags_type']

        # Validation du format du numéro de camion
        if not validate_truck_number(truck_number):
            flash('Invalid truck number format! Please use one of the following formats: A1245RB, AA1234RB, A3345RB/A4567RB, etc.', 'error')
            return redirect(url_for('edit_container', id=id))

        # Validation du format du numéro de conteneur
        if not validate_container_number(container_name):
            flash('Invalid container name format! Please use the format: ABCD1234567.', 'error')
            return redirect(url_for('edit_container', id=id))

        # Vérification de l'unicité du numéro de conteneur
        existing_container = Container.query.filter_by(container_name=container_name).first()
        if existing_container and existing_container.id != id:
            flash('Container number already exists! Please use a different container number.', 'error')
            return redirect(url_for('edit_container', id=id))

        # Met à jour le conteneur avec les nouvelles informations
        container.truck_number = truck_number
        container.container_name = container_name
        container.plant = plant
        container.product = product
        container.freight_forwarder = freight_forwarder
        container.booking_id = booking_id
        container.container_tare = container_tare
        container.bags_type = bags_type
        container.container_created_by = current_user.username

        # Enregistre les modifications dans la base de données
        db.session.commit()
        flash('Container updated successfully!', 'success')
        return redirect(url_for('containers'))

    # Affiche le formulaire de modification de conteneur avec la liste des réservations
    bookings = Booking.query.all()
    return render_template('container/edit_container.html', container=container, bookings=bookings)

@app.route('/container/<int:id>/delete', methods=['POST'])
@login_required
@role_required('System Administrator')
def delete_container(id):
    """
    Permet de supprimer un conteneur spécifique. 
    Accessible uniquement aux utilisateurs authentifiés ayant le rôle approprié.
    """
    container = Container.query.get_or_404(id)
    
    # Vérifie les associations avec la table Loading
    loadings = Loading.query.filter_by(container_id=id).all()
    if loadings:
        flash(f'Cannot delete Container {container.container_name} because it is associated with loading records.', 'error')
        return redirect(url_for('containers'))

    # Vérifie les associations avec la table Seal
    seals = Seal.query.filter_by(container_id=id).all()
    if seals:
        flash(f'Cannot delete Container {container.container_name} because it is associated with seal records.', 'error')
        return redirect(url_for('containers'))

    # Vérifie les associations avec la table Weight
    weights = Weight.query.filter_by(container_id=id).all()
    if weights:
        flash(f'Cannot delete Container {container.container_name} because it is associated with weight records.', 'error')
        return redirect(url_for('containers'))

    # Si aucune association n'est trouvée, procède à la suppression
    db.session.delete(container)
    db.session.commit()
    flash(f'Container {container.container_name} deleted successfully', 'success')

    return redirect(url_for('containers'))

@app.route('/containers')
#@login_required
#@role_required('Logistics Manager', 'Inventory Manager', 'Purchasing Manager', 'Transportation Coordinator', 'Warehouse Operator', 'Manager', 'Logistics Data Analyst')
def containers():
    """
    Affiche la liste de tous les conteneurs.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    containers = Container.query.all()

    # Convertit les objets Container en dictionnaires pour l'affichage
    containers_list = [
        {
            'id': container.id,
            'arrival_date': container.arrival_date.strftime('%Y-%m-%d %H:%M:%S'),
            'truck_number': container.truck_number,
            'container_name': container.container_name,
            'plant': container.plant,
            'product': container.product,
            'freight_forwarder': container.freight_forwarder,
            'booking_id': container.booking_id,
            'booking_name': container.booking.booking_name if container.booking else 'Unknown',
            'container_tare': container.container_tare,
            'bags_type': container.bags_type,
            'container_created_at': container.container_created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'created_by': container.container_created_by,
            'status': container.get_status()
        }
        for container in containers
    ]

    return render_template('container/containers.html', containers=containers_list)

@app.route('/container/<int:id>/view')
@login_required
@role_required('Logistics Manager', 'Inventory Manager', 'Purchasing Manager', 'Transportation Coordinator', 'Warehouse Operator', 'Manager', 'Customer Support Representative')
def view_container(id):
    """
    Affiche les détails d'un conteneur spécifique, y compris les informations associées.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    container = Container.query.get_or_404(id)
    loading = Loading.query.filter_by(container_id=id).first()
    seal = Seal.query.filter_by(container_id=id).first()
    weight = Weight.query.filter_by(container_id=id).first()

    return render_template('container/view_container.html', 
                           container=container, 
                           loading=loading, 
                           seal=seal, 
                           weight=weight)

#------------------------------Add Loading fr contrainer ------------------------------------------


temp_loading_data = None

@login_required
@app.route('/loading/add/<int:container_id>', methods=['GET', 'POST'])
@role_required('Logistics Manager', 'Warehouse Operator', 'Transportation Coordinator', 'Manager')
def add_loading(container_id):
    """
    Permet d'ajouter un nouveau chargement pour un conteneur spécifique.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    global temp_loading_data

    if request.method == 'POST':
        # Sauvegarde temporaire des données de chargement
        temp_loading_data = {
            'container_id': container_id,
            'no_of_bags': request.form.get('no_of_bags'),
            'labor': request.form.get('labor'),
            'loading_created_by': current_user.username
        }
        return redirect(url_for('confirm_loading_creation'))

    return render_template('loading/create_loading.html', container_id=container_id)

@login_required
@app.route('/confirm_loading_creation', methods=['GET', 'POST'])
@role_required('Logistics Manager', 'Warehouse Operator', 'Transportation Coordinator', 'Manager')
def confirm_loading_creation():
    """
    Confirme la création d'un nouveau chargement basé sur les données temporaires.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    global temp_loading_data

    if request.method == 'POST':
        confirm_choice = request.form.get('choice')

        if confirm_choice == 'confirm':
            # Crée un nouvel objet Loading et l'ajoute à la base de données
            new_loading = Loading(
                container_id=temp_loading_data['container_id'],
                no_of_bags=temp_loading_data['no_of_bags'],
                labor=temp_loading_data['labor'],
                loading_created_by=current_user.username
            )
            db.session.add(new_loading)
            db.session.commit()
            flash('Loading added successfully', 'success')
        elif confirm_choice == 'cancel':
            flash('Loading addition canceled.', 'info')

        temp_loading_data = None
        return redirect(url_for('containers'))

    if temp_loading_data:
        loading_data = temp_loading_data
    else:
        return redirect(url_for('add_loading'))

    return render_template('loading/confirm_loading_creation.html', loading_data=loading_data)

@login_required
@app.route('/loading/<int:id>/edit', methods=['GET', 'POST'])
@role_required('Logistics Manager', 'Warehouse Operator', 'Transportation Coordinator', 'Manager')
def edit_loading(id):
    """
    Permet de modifier les détails d'un chargement spécifique.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    loading = Loading.query.get_or_404(id)

    if request.method == 'POST':
        # Met à jour les attributs de l'objet Loading
        loading.no_of_bags = request.form['no_of_bags']
        loading.labor = request.form['labor']
        loading.loading_created_by = current_user.username
        db.session.commit()
        flash('Loading updated successfully', 'success')
        return redirect(url_for('containers'))

    return render_template('loading/edit_loading.html',id=loading.id, loading=loading)

@login_required
@app.route('/loading/<int:id>/delete', methods=['POST'])
@role_required('System Administrator')
def delete_loading(id):
    """
    Supprime un chargement spécifique.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    loading = Loading.query.get_or_404(id)
    db.session.delete(loading)
    db.session.commit()
    flash('Loading deleted successfully', 'success')
    return redirect(url_for('containers'))

@login_required
@app.route('/loading/<int:id>')
@role_required('Logistics Manager', 'Inventory Manager', 'Warehouse Operator', 'Transportation Coordinator', 'Manager', 'Customer Support Representative')
def view_loading(id):
    """
    Affiche les détails d'un chargement spécifique.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    loading = Loading.query.get_or_404(id)
    return render_template('loading/view_loading.html',id=loading.id, loading=loading)

#------------------------------Seal Management ------------------------

temp_seal_data = None

@login_required
@app.route('/seal/add/<int:container_id>', methods=['GET', 'POST'])
@role_required('Logistics Manager', 'Warehouse Operator', 'Transportation Coordinator', 'Manager')
def add_seal(container_id):
    """
    Permet d'ajouter un nouveau sceau pour un conteneur spécifique.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    global temp_seal_data
    if request.method == 'POST':
        seal_number = request.form['seal_number']
        seal_date = request.form['seal_date']
        seal_created_by = current_user.username
        seal_image = request.files['seal_image']
        image_filename = None
        
        if seal_image:
            image_filename = f"{seal_number}.jpg"
            seal_image.save(os.path.join(app.config['UPLOAD_FOLDER_SEAL'], image_filename))

        temp_seal_data = {
            'seal_number': seal_number,
            'seal_image': image_filename,
            'seal_date': seal_date,
            'seal_created_by': seal_created_by,
            'container_id': container_id
        }

        return redirect(url_for('confirm_create_seal'))
    return render_template('seal/add_seal.html', container_id=container_id)

@login_required
@app.route('/confirm_create_seal', methods=['GET', 'POST'])
@role_required('Logistics Manager', 'Warehouse Operator', 'Transportation Coordinator', 'Manager')
def confirm_create_seal():
    """
    Confirme la création d'un nouveau sceau basé sur les données temporaires.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    global temp_seal_data

    if request.method == 'POST':
        confirm_choice = request.form.get('choice')

        if confirm_choice == 'confirm':
            new_seal = Seal(
                seal_number=temp_seal_data['seal_number'],
                seal_image=temp_seal_data['seal_image'],
                seal_date=datetime.strptime(temp_seal_data['seal_date'], '%Y-%m-%d').date(),
                seal_created_by=current_user.username,
                container_id=temp_seal_data['container_id']
            )

            db.session.add(new_seal)
            db.session.commit()
            flash('Seal added successfully', 'success')
        elif confirm_choice == 'cancel':
            flash('Seal addition canceled.', 'info')

        temp_seal_data = None
        return redirect(url_for('seals'))

    if temp_seal_data:
        return render_template('seal/confirm_seal_creation.html', seal_data=temp_seal_data)
    else:
        return redirect(url_for('add_seal', container_id=temp_seal_data['container_id']))

@login_required
@app.route('/seal/<int:id>/edit', methods=['GET', 'POST'])
@role_required('Logistics Manager', 'Warehouse Operator', 'Transportation Coordinator', 'Manager')
def edit_seal(id):
    """
    Permet de modifier les détails d'un sceau spécifique.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    seal = Seal.query.get_or_404(id)

    if request.method == 'POST':
        seal.seal_number = request.form['seal_number']
        seal.seal_date = datetime.strptime(request.form['seal_date'], '%Y-%m-%d')
        seal.seal_created_by = current_user.username

        new_seal_image = request.files['seal_image']
        if new_seal_image:
            image_filename = f"{seal.seal_number}.jpg"
            new_seal_image.save(os.path.join(app.config['UPLOAD_FOLDER_SEAL'], image_filename))
            seal.seal_image = image_filename

        db.session.commit()
        flash('Seal updated successfully', 'success')
        return redirect(url_for('seals'))

    return render_template('seal/edit_seal.html', seal=seal)

@login_required
@app.route('/seal/<int:id>/delete', methods=['POST'])
@role_required('System Administrator')
def delete_seal(id):
    """
    Supprime un sceau spécifique.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    seal = Seal.query.get_or_404(id)
    db.session.delete(seal)
    db.session.commit()
    flash('Seal deleted successfully', 'success')
    return redirect(url_for('seals'))

@app.route('/seal/<int:id>')
@login_required
@role_required('Logistics Manager', 'Inventory Manager', 'Warehouse Operator', 'Transportation Coordinator', 'Manager', 'Customer Support Representative')
def view_seal(id):
    """
    Affiche les détails d'un sceau spécifique.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    seal = Seal.query.get(id)
    return render_template('seal/view_seal.html', seal=seal)

@app.route('/seals')
@login_required
def seals():
    """
    Affiche la liste de tous les sceaux.
    Accessible uniquement aux utilisateurs authentifiés.
    """
    seals = Seal.query.all()
    seals_list = [
        {
            'id': seal.id,
            'seal_number': seal.seal_number,
            'seal_image': seal.seal_image,
            'seal_date': seal.seal_date.strftime('%Y-%m-%d'),
            'seal_created_by': seal.seal_created_by,
            'container_id': seal.container_id,
            'container_name': seal.container.container_name if seal.container else 'N/A'  # Ajout de l'information du conteneur
        }
        for seal in seals
    ]
    return render_template('seal/seals.html', seals=seals_list)




#--------------------------Weight Management --------------------------------



temp_weight_data = None

@login_required
@app.route('/weight/add/<int:container_id>', methods=['GET', 'POST'])
@role_required('Logistics Manager', 'Warehouse Operator', 'Transportation Coordinator', 'Manager')
def add_weight(container_id):
    """
    Permet d'ajouter un nouveau poids pour un conteneur spécifique.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    global temp_weight_data
    if request.method == 'POST':
        ws_number = request.form['ws_number']
        gross_weight = request.form['gross_weight']
        tare_weight = request.form['tare_weight']
        net_weight = request.form['net_weight']
        weight_date = request.form['weight_date']
        weight_created_by = current_user.username
        ws_image = request.files['ws_image']
        image_filename = None
        
        if ws_image:
            image_filename = f"{container_id}_{ws_number}.jpg"
            ws_image.save(os.path.join(app.config['UPLOAD_FOLDER_WEIGHT'], image_filename))

        temp_weight_data = {
            'ws_number': ws_number,
            'ws_image': image_filename,
            'gross_weight': gross_weight,
            'tare_weight': tare_weight,
            'net_weight': net_weight,
            'weight_date': weight_date,
            'weight_created_by': weight_created_by,
            'container_id': container_id
        }

        return redirect(url_for('confirm_create_weight'))

    return render_template('weight/add_weight.html', container_id=container_id, )

@login_required
@app.route('/confirm_create_weight', methods=['GET', 'POST'])
@role_required('Logistics Manager', 'Warehouse Operator', 'Transportation Coordinator', 'Manager')
def confirm_create_weight():
    """
    Confirme la création d'un nouveau poids basé sur les données temporaires.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    global temp_weight_data

    if request.method == 'POST':
        confirm_choice = request.form.get('choice')

        if confirm_choice == 'confirm':
            new_weight = Weight(
                ws_number=temp_weight_data['ws_number'],
                ws_image=temp_weight_data['ws_image'],
                gross_weight=int(temp_weight_data['gross_weight']),
                tare_weight=int(temp_weight_data['tare_weight']),
                net_weight=int(temp_weight_data['net_weight']),
                weight_date=datetime.strptime(temp_weight_data['weight_date'], '%Y-%m-%d').date(),
                weight_created_by=current_user.username,
                container_id=temp_weight_data['container_id']
            )

            db.session.add(new_weight)
            db.session.commit()
            flash('Weight added successfully', 'success')
        elif confirm_choice == 'cancel':
            flash('Weight addition canceled.', 'info')

        temp_weight_data = None
        return redirect(url_for('weights'))

    if temp_weight_data:
        return render_template('weight/confirm_weight_creation.html', weight_data=temp_weight_data)
    else:
        return redirect(url_for('add_weight'))

@login_required
@app.route('/weight/<int:id>/edit', methods=['GET', 'POST'])
@role_required('Logistics Manager', 'Warehouse Operator', 'Transportation Coordinator', 'Manager')
def edit_weight(id):
    """
    Permet de modifier les détails d'un poids spécifique.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    weight = Weight.query.get_or_404(id)

    if request.method == 'POST':
        weight.ws_number = request.form['ws_number']
        weight.gross_weight = request.form['gross_weight']
        weight.tare_weight = request.form['tare_weight']
        weight.net_weight = request.form['net_weight']
        weight.weight_date = datetime.strptime(request.form['weight_date'], '%Y-%m-%d')
        weight.weight_created_by = current_user.username

        new_ws_image = request.files['ws_image']
        if new_ws_image:
            image_filename = f"{weight.container_id}_{weight.ws_number}.jpg"
            new_ws_image.save(os.path.join(app.config['UPLOAD_FOLDER_WEIGHT'], image_filename))
            weight.ws_image = image_filename

        db.session.commit()
        flash('Weight updated successfully', 'success')
        return redirect(url_for('weights'))

    return render_template('weight/edit_weight.html', weight=weight)

@login_required
@app.route('/weight/<int:id>/delete', methods=['POST'])
@role_required('System Administrator')
def delete_weight(id):
    """
    Supprime un poids spécifique.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    weight = Weight.query.get_or_404(id)
    db.session.delete(weight)
    db.session.commit()
    flash('Weight deleted successfully', 'success')
    return redirect(url_for('weights'))

@app.route('/weight/<int:id>')
@login_required
@role_required('Logistics Manager', 'Inventory Manager', 'Warehouse Operator', 'Transportation Coordinator', 'Manager', 'Customer Support Representative')
def view_weight(id):
    """
    Affiche les détails d'un poids spécifique.
    Accessible uniquement aux utilisateurs authentifiés ayant les bons rôles.
    """
    weight = Weight.query.get(id)
    if weight is None:
        return "Weight record not found", 404
    return render_template('weight/view_weight.html', weight=weight)

@app.route('/weights')
@login_required
def weights():
    """
    Affiche la liste de tous les poids.
    Accessible uniquement aux utilisateurs authentifiés.
    """
    weights = Weight.query.all()
    
    weights_list = [
        {
            'id': weight.id,
            'ws_number': weight.ws_number,
            'ws_image': weight.ws_image,
            'gross_weight': weight.gross_weight,
            'tare_weight': weight.tare_weight,
            'net_weight': weight.net_weight,
            'weight_date': weight.weight_date.strftime('%Y-%m-%d') if weight.weight_date else 'None',
            'weight_created_by': weight.weight_created_by,
            'container_id': weight.container_id,
            'container_number': weight.container.container_name if weight.container else 'None'
        }
        for weight in weights
    ]
    
    return render_template('weight/weights.html', weights=weights_list)



# Variable globale pour stocker temporairement les données de paiement
temp_payment_info_data = None

# Route pour ajouter des informations de paiement
@app.route('/contract/<int:contract_id>/add_payment_info', methods=['GET', 'POST'])
@login_required
@role_required('Logistics Manager', 'Purchasing Manager', 'System Administrator', 'Manager', 'Financial Analyst')
def add_payment_info(contract_id):
    contract = Contract.query.get_or_404(contract_id)
    
    if contract.payment_info:
        flash('Payment information already exists for this contract.', 'warning')
        return redirect(url_for('view_contract', id=contract_id))
    
    if request.method == 'POST':
        payment_term = request.form['payment_term']
        price = request.form['price']
        
        global temp_payment_info_data
        temp_payment_info_data = {
            'payment_term': payment_term,
            'price': price,
            'contract_id': contract_id
        }
        
        return redirect(url_for('confirm_payment_info_creation'))
    
    return render_template('payment_info/add_payment_info.html', contract=contract)

# Route pour confirmer la création des informations de paiement
@app.route('/confirm_payment_info_creation', methods=['GET', 'POST'])
@login_required
@role_required('Logistics Manager', 'Purchasing Manager', 'System Administrator', 'Manager', 'Financial Analyst')
def confirm_payment_info_creation():
    global temp_payment_info_data

    if request.method == 'POST':
        confirm_choice = request.form.get('choice')

        if confirm_choice == 'confirm':
            new_payment_info = PaymentInfo(
                contract_id=temp_payment_info_data['contract_id'],
                payment_term=temp_payment_info_data['payment_term'],
                price=temp_payment_info_data['price']
            )

            db.session.add(new_payment_info)
            db.session.commit()

            flash('Payment information added successfully', 'success')
        elif confirm_choice == 'cancel':
            flash('Payment information addition canceled.', 'info')

        temp_payment_info_data = None

        return redirect(url_for('contracts'))

    if temp_payment_info_data:
        payment_info_data = temp_payment_info_data
        contract = Contract.query.get(payment_info_data['contract_id'])
        payment_info_data['contract_number'] = contract.contract_number if contract else 'Unknown'
    else:
        return redirect(url_for('add_payment_info'))

    return render_template('payment_info/confirm_payment_info_creation.html', payment_info_data=payment_info_data)

# Route pour éditer les informations de paiement
@app.route('/payment_info/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('Logistics Manager', 'Purchasing Manager', 'System Administrator', 'Manager', 'Financial Analyst')
def edit_payment_info(id):
    payment_info = PaymentInfo.query.get_or_404(id)

    if request.method == 'POST':
        payment_term = request.form['payment_term']
        price = request.form['price']

        payment_info.payment_term = payment_term
        payment_info.price = price

        db.session.commit()
        flash('Payment information updated successfully!', 'success')
        return redirect(url_for('view_contract', id=payment_info.contract_id))

    return render_template('payment_info/edit_payment_info.html', payment_info=payment_info)

# Route pour supprimer les informations de paiement
@app.route('/payment_info/<int:id>/delete', methods=['POST'])
@login_required
@role_required('System Administrator')
def delete_payment_info(id):
    payment_info = PaymentInfo.query.get_or_404(id)
    db.session.delete(payment_info)
    db.session.commit()
    flash('Payment information deleted successfully', 'success')
    return redirect(url_for('view_contract', id=payment_info.contract_id))



# Fonction utilitaire pour convertir une chaîne en date
def convert_to_date(date_str):
    if date_str:
        try:
            return datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return None
    return None

# Variable globale pour stocker temporairement les données de date
temp_date_info_data = None

# Route pour ajouter des informations de date
@app.route('/contract/<int:contract_id>/add_date_info', methods=['GET', 'POST'])
@login_required
@role_required('Logistics Manager', 'Purchasing Manager', 'System Administrator', 'Manager', 'Financial Analyst')
def add_date_info(contract_id):
    contract = Contract.query.get_or_404(contract_id)
    
    if contract.date_info:
        flash('Date information already exists for this contract.', 'warning')
        return redirect(url_for('view_contract', id=contract_id))
    
    if request.method == 'POST':
        contract_copy_date = request.form['contract_copy_date']
        signed_contract_date = request.form['signed_contract_date']
        si1_date = request.form['si1_date']
        si2_date = request.form['si2_date']
        si3_date = request.form['si3_date']
        lc_date = request.form['lc_date']
        
        global temp_date_info_data
        temp_date_info_data = {
            'contract_copy_date': contract_copy_date,
            'signed_contract_date': signed_contract_date,
            'si1_date': si1_date,
            'si2_date': si2_date,
            'si3_date': si3_date,
            'lc_date': lc_date,
            'contract_id': contract_id
        }
        
        return redirect(url_for('confirm_date_info_creation'))
    
    return render_template('date_info/add_date_info.html', contract=contract)

# Route pour confirmer la création des informations de date
@app.route('/confirm_date_info_creation', methods=['GET', 'POST'])
@login_required
@role_required('Logistics Manager', 'Purchasing Manager', 'System Administrator', 'Manager', 'Financial Analyst')
def confirm_date_info_creation():
    global temp_date_info_data

    if request.method == 'POST':
        confirm_choice = request.form.get('choice')

        if confirm_choice == 'confirm':
            new_date_info = DateInfo(
                contract_id=temp_date_info_data['contract_id'],
                contract_copy_date=convert_to_date(temp_date_info_data['contract_copy_date']),
                signed_contract_date=convert_to_date(temp_date_info_data['signed_contract_date']),
                si1_date=convert_to_date(temp_date_info_data['si1_date']),
                si2_date=convert_to_date(temp_date_info_data['si2_date']),
                si3_date=convert_to_date(temp_date_info_data['si3_date']),
                lc_date=convert_to_date(temp_date_info_data['lc_date'])
            )

            db.session.add(new_date_info)
            db.session.commit()

            flash('Date information added successfully', 'success')
        elif confirm_choice == 'cancel':
            flash('Date information addition canceled.', 'info')

        temp_date_info_data = None

        return redirect(url_for('contracts'))

    if temp_date_info_data:
        date_info_data = temp_date_info_data
        contract = Contract.query.get(date_info_data['contract_id'])
        date_info_data['contract_number'] = contract.contract_number if contract else 'Unknown'
    else:
        return redirect(url_for('add_date_info'))

    return render_template('date_info/confirm_date_info_creation.html', date_info_data=date_info_data)

# Route pour éditer les informations de date
@app.route('/date_info/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('Logistics Manager', 'Purchasing Manager', 'System Administrator', 'Manager', 'Financial Analyst')
def edit_date_info(id):
    date_info = DateInfo.query.get_or_404(id)

    if request.method == 'POST':
        contract_copy_date = request.form['contract_copy_date']
        signed_contract_date = request.form['signed_contract_date']
        si1_date = request.form['si1_date']
        si2_date = request.form['si2_date']
        si3_date = request.form['si3_date']
        lc_date = request.form['lc_date']

        date_info.contract_copy_date = convert_to_date(contract_copy_date)
        date_info.signed_contract_date = convert_to_date(signed_contract_date)
        date_info.si1_date = convert_to_date(si1_date)
        date_info.si2_date = convert_to_date(si2_date)
        date_info.si3_date = convert_to_date(si3_date)
        date_info.lc_date = convert_to_date(lc_date)

        db.session.commit()
        flash('Date information updated successfully!', 'success')
        return redirect(url_for('view_contract', id=date_info.contract_id))

    return render_template('date_info/edit_date_info.html', date_info=date_info)

# Route pour supprimer les informations de date
@app.route('/date_info/<int:id>/delete', methods=['POST'])
@login_required
@role_required('System Administrator')
def delete_date_info(id):
    date_info = DateInfo.query.get_or_404(id)
    db.session.delete(date_info)
    db.session.commit()
    flash('Date information deleted successfully', 'success')
    return redirect(url_for('view_contract', id=date_info.contract_id))

#---------- Remarks --------------------------------#
# Variable globale pour stocker temporairement les données de remarques
temp_remarks_info_data = None

# Route pour ajouter des informations de remarques
@app.route('/contract/<int:contract_id>/add_remarks_info', methods=['GET', 'POST'])
@login_required
@role_required('Logistics Manager', 'Purchasing Manager', 'System Administrator', 'Manager', 'Financial Analyst')
def add_remarks_info(contract_id):
    contract = Contract.query.get_or_404(contract_id)
    
    if contract.remarks_info:
        flash('Remarks information already exists for this contract.', 'warning')
        return redirect(url_for('view_contract', id=contract_id))
    
    if request.method == 'POST':
        first_remarks = request.form['first_remarks']
        second_remarks = request.form['second_remarks']
        
        global temp_remarks_info_data
        temp_remarks_info_data = {
            'first_remarks': first_remarks,
            'second_remarks': second_remarks,
            'contract_id': contract_id
        }
        
        return redirect(url_for('confirm_remarks_info_creation'))
    
    return render_template('remarks_info/add_remarks_info.html', contract=contract)

# Route pour confirmer la création des informations de remarques
@app.route('/confirm_remarks_info_creation', methods=['GET', 'POST'])
@login_required
@role_required('Logistics Manager', 'Purchasing Manager', 'System Administrator', 'Manager', 'Financial Analyst')
def confirm_remarks_info_creation():
    global temp_remarks_info_data

    if request.method == 'POST':
        confirm_choice = request.form.get('choice')

        if confirm_choice == 'confirm':
            new_remarks_info = RemarksInfo(
                contract_id=temp_remarks_info_data['contract_id'],
                first_remarks=temp_remarks_info_data['first_remarks'],
                second_remarks=temp_remarks_info_data['second_remarks']
            )

            db.session.add(new_remarks_info)
            db.session.commit()

            flash('Remarks information added successfully', 'success')
        elif confirm_choice == 'cancel':
            flash('Remarks information addition canceled.', 'info')

        temp_remarks_info_data = None

        return redirect(url_for('contracts'))

    if temp_remarks_info_data:
        remarks_info_data = temp_remarks_info_data
        contract = Contract.query.get(remarks_info_data['contract_id'])
        remarks_info_data['contract_number'] = contract.contract_number if contract else 'Unknown'
    else:
        return redirect(url_for('add_remarks_info'))

    return render_template('remarks_info/confirm_remarks_info_creation.html', remarks_info_data=remarks_info_data)

# Route pour éditer les informations de remarques
@app.route('/remarks_info/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('Logistics Manager', 'Purchasing Manager', 'System Administrator', 'Manager', 'Financial Analyst','Finance Manager')
def edit_remarks_info(id):
    remarks_info = RemarksInfo.query.get_or_404(id)

    if request.method == 'POST':
        first_remarks = request.form['first_remarks']
        second_remarks = request.form['second_remarks']

        remarks_info.first_remarks = first_remarks
        remarks_info.second_remarks = second_remarks

        db.session.commit()
        flash('Remarks information updated successfully!', 'success')
        return redirect(url_for('view_contract', id=remarks_info.contract_id))

    return render_template('remarks_info/edit_remarks_info.html', remarks_info=remarks_info)

# Route pour supprimer les informations de remarques
@app.route('/remarks_info/<int:id>/delete', methods=['POST'])
@login_required
@role_required('System Administrator')
def delete_remarks_info(id):
    remarks_info = RemarksInfo.query.get_or_404(id)
    db.session.delete(remarks_info)
    db.session.commit()
    flash('Remarks information deleted successfully', 'success')
    return redirect(url_for('view_contract', id=remarks_info.contract_id))


#-------------------------------Admin Interface Function---------------------------------------------

@app.route('/export-excel')
@role_required('System Administrator', 'Logistics Manager', 'Logistics Data Analyst','Finance','Finance Manager')
def export_excel():
    # Créez un objet BytesIO pour stocker le fichier Excel en mémoire
    output = BytesIO()
    
    # Récupérer toutes les données
    containers = Container.query.all()
    
    # Préparez une liste pour stocker les données combinées
    combined_data = []

    for container in containers:
        # Obtenez les données associées
        booking = Booking.query.get(container.booking_id)
        contract = Contract.query.get(booking.contract_id) if booking else None
        customer = Customer.query.get(contract.customer_id) if contract else None
        weight = Weight.query.filter_by(container_id=container.id).first()
        seal = Seal.query.filter_by(container_id=container.id).first()
        loading = Loading.query.filter_by(container_id=container.id).first()
        
        # Combinez les données en un dictionnaire
        combined_data.append({
            'Container ID': container.id,
            'Container Name': container.container_name,
            'Arrival Date': container.arrival_date,
            'Truck Number': container.truck_number,
            'Plant': container.plant,
            'Product': container.product,
            'Freight Forwarder': container.freight_forwarder,
            'Container Tare': container.container_tare,
            'Bags Type': container.bags_type,
            'Booking Name': booking.booking_name if booking else 'None',
            'Contract Number': contract.contract_number if contract else 'None',
            'Customer Name': customer.name if customer else 'None',
            'Customer Email': customer.email if customer else 'None',
            'Net Weight': weight.net_weight if weight else 'None',
            'Weight Date': weight.weight_date if weight else 'None',
            'Seal Number': seal.seal_number if seal else 'None',
            'Seal Date': seal.seal_date if seal else 'None',
            'Loading No of Bags': loading.no_of_bags if loading else 'None',
            'Loading Labor': loading.labor if loading else 'None',
            'Loading Created At': loading.loading_created_at if loading else 'None',
        })
    
    # Convertir les données combinées en DataFrame
    df = pd.DataFrame(combined_data)
    
    # Créer un nouveau classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = 'Containers Report'
    
    # Écrire les en-têtes dans la première ligne avec mise en forme
    headers = df.columns.tolist()
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")  # Texte en gras et blanc
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Couleur d'arrière-plan bleu
        cell.alignment = Alignment(horizontal='center', vertical='center')  # Alignement centré

    # Écrire les données avec mise en forme
    for row_num, row_data in enumerate(df.values.tolist(), 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            cell.alignment = Alignment(horizontal='left', vertical='center')  # Alignement des cellules

    # Ajouter un tableau à la feuille de calcul
    table = Table(displayName="ContainersTable", ref=f"A1:{ws.cell(row=1, column=len(headers)).column_letter}{len(df) + 1}")

    # Ajouter un style au tableau
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    table.tableStyleInfo = style

    # Ajouter le tableau à la feuille de calcul
    ws.add_table(table)

    # Ajuster la largeur des colonnes en fonction du contenu
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Obtenir le nom de la colonne
        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except:  # noqa: E722
                pass
        adjusted_width = max_length + 2  # Ajouter un peu d'espace pour la lisibilité
        ws.column_dimensions[column].width = adjusted_width

    # Enregistrer le fichier Excel en mémoire
    wb.save(output)
    output.seek(0)
    
    # Obtenir la date actuelle et formater le nom du fichier
    current_date = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    file_name = f'containers_report_{current_date}.xlsx'
    
    # Retourner le fichier Excel avec le nom de fichier formaté
    return send_file(output, as_attachment=True, download_name=file_name, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# Assurez-vous d'avoir la configuration de Flask et les modèles définis ici...


@app.route('/report-container')
@role_required('System Administrator', 'Logistics Manager', 'Logistics Data Analyst','Finance Manager')
def report_container():
    containers = Container.query.all()
    data = []
    for container in containers:
        booking = Booking.query.get(container.booking_id)
        contract = Contract.query.get(booking.contract_id) if booking else None
        
        # Récupérer les informations de loading, seal et weight
        loading = container.loading
        seal = container.seal
        weight = container.weight

        data.append({
            'container_id': container.id,
            'loading_date': loading.loading_created_at.strftime('%Y-%m-%d %H:%M:%S') if loading else 'None',
            'container_name': container.container_name,
            'product': container.product,
            'plant': container.plant,
            'bags': loading.no_of_bags if loading else 'None',
            'booking_name': booking.booking_name if booking else 'None',
            'contract_number': contract.contract_number if contract else 'None',
            'seal_number': seal.seal_number if seal else 'None',
            'seal_date': seal.seal_date if seal else 'None',
            'gross_weight': weight.gross_weight if weight else 'None',
            'tare_weight': weight.tare_weight if weight else 'None',
            'net_weight': weight.net_weight if weight else 'None',
            'weight_date': weight.weight_date if weight else 'None'
        })
    return render_template('container/report_container.html', data=data)


@app.route('/generate-excel-report/<int:container_id>', methods=['GET'])
@role_required('System Administrator', 'Logistics Manager', 'Logistics Data Analyst','Finance')
def generate_excel_report(container_id):
    # Créez un objet BytesIO pour stocker le fichier Excel en mémoire
    output = BytesIO()

    # Récupérer les données pour le conteneur spécifique
    container = Container.query.get(container_id)
    if not container:
        abort(404, description="Container not found")

    # Obtenez les données associées
    booking = Booking.query.get(container.booking_id)
    contract = Contract.query.get(booking.contract_id) if booking else None
    customer = Customer.query.get(contract.customer_id) if contract else None
    weight = Weight.query.filter_by(container_id=container.id).first()
    seal = Seal.query.filter_by(container_id=container.id).first()
    loading = Loading.query.filter_by(container_id=container.id).first()

    # Combinez les données en un dictionnaire
    combined_data = [{
        'Container ID': container.id,
        'Container Name': container.container_name,
        'Arrival Date': container.arrival_date,
        'Truck Number': container.truck_number,
        'Plant': container.plant,
        'Product': container.product,
        'Freight Forwarder': container.freight_forwarder,
        'Container Tare': container.container_tare,
        'Bags Type': container.bags_type,
        'Booking Name': booking.booking_name if booking else 'None',
        'Contract Number': contract.contract_number if contract else 'None',
        'Customer Name': customer.name if customer else 'None',
        'Customer Email': customer.email if customer else 'None',
        'Net Weight': weight.net_weight if weight else 'None',
        'Weight Date': weight.weight_date if weight else 'None',
        'Seal Number': seal.seal_number if seal else 'None',
        'Seal Date': seal.seal_date if seal else 'None',
        'Loading No of Bags': loading.no_of_bags if loading else 'None',
        'Loading Labor': loading.labor if loading else 'None',
        'Loading Created At': loading.loading_created_at if loading else 'None',
    }]
    
    # Convertir les données en DataFrame
    df = pd.DataFrame(combined_data)
    
    # Créer un nouveau classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = 'Container Report'
    
    # Écrire les en-têtes dans la première ligne avec mise en forme
    headers = df.columns.tolist()
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")  # Texte en gras et blanc
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Couleur d'arrière-plan bleu
        cell.alignment = Alignment(horizontal='center', vertical='center')  # Alignement centré

    # Écrire les données avec mise en forme
    for row_num, row_data in enumerate(df.values.tolist(), 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            cell.alignment = Alignment(horizontal='left', vertical='center')  # Alignement des cellules

    # Ajouter un tableau à la feuille de calcul
    table = Table(displayName="ContainerTable", ref=f"A1:{ws.cell(row=1, column=len(headers)).column_letter}{len(df) + 1}")

    # Ajouter un style au tableau
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    table.tableStyleInfo = style

    # Ajouter le tableau à la feuille de calcul
    ws.add_table(table)

    # Ajuster la largeur des colonnes en fonction du contenu
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Obtenir le nom de la colonne
        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except:  # noqa: E722
                pass
        adjusted_width = max_length + 2  # Ajouter un peu d'espace pour la lisibilité
        ws.column_dimensions[column].width = adjusted_width

    # Enregistrer le fichier Excel en mémoire
    wb.save(output)
    output.seek(0)
    
    # Retourner le fichier Excel
    return send_file(output, as_attachment=True, download_name=f'Container_{container.container_name}_Report.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# Définir le filtre personnalisé
@app.template_filter('format_number')
def format_number(value):
    try:
        # Assurez-vous que la valeur est un nombre flottant
        num = float(value)
        return '{:.0f}'.format(num) if num % 1 == 0 else '{:.2f}'.format(num).rstrip('0').rstrip('.')
    except (ValueError, TypeError):
        return value

@app.route('/generate_pdf_report/<int:container_id>')
@role_required('System Administrator','Finance Manager')
def generate_pdf_report(container_id):
    # Récupérer les données du conteneur
    container = Container.query.get(container_id)
    if not container:
        abort(404, description="Container not found")

    # Récupérer les données associées
    booking = Booking.query.get(container.booking_id)
    contract = Contract.query.get(booking.contract_id) if booking else None
    weight = Weight.query.filter_by(container_id=container_id).first()
    seal = Seal.query.filter_by(container_id=container_id).first()
    loading = Loading.query.filter_by(container_id=container_id).first()

    # Récupérer les utilisateurs qui ont créé les enregistrements
    container_creator = User.query.filter_by(username=container.container_created_by).first()
    loading_creator = User.query.filter_by(username=loading.loading_created_by).first() if loading else None
    seal_creator = User.query.filter_by(username=seal.seal_created_by).first() if seal else None
    weight_creator = User.query.filter_by(username=weight.weight_created_by).first() if weight else None
    
    # Obtenir les noms complets des créateurs
    container_creator_full_name = container_creator.full_name if container_creator else 'N/A'
    loading_creator_full_name = loading_creator.full_name if loading_creator else 'N/A'
    seal_creator_full_name = seal_creator.full_name if seal_creator else 'N/A'
    weight_creator_full_name = weight_creator.full_name if weight_creator else 'N/A'

    # Choisir le logo en fonction de la valeur de plant
    if contract and contract.plant == 'BO':
        logo_filename = 'logo_bo.png'
    elif contract and contract.plant == 'BAB':
        logo_filename = 'logo_bab.png'
    else:
        logo_filename = 'default_logo.png'  # Un logo par défaut si nécessaire

    # Construire l'URL du logo
    logo_url = url_for('static', filename=logo_filename, _external=True)

    # Préparer les données pour le template
    data = {
        'image_url': logo_url,
        'booking_name': booking.booking_name if booking else 'N/A',
        'container_name': container.container_name,
        'arrival_date': container.arrival_date.strftime('%d-%m-%Y'),
        'truck_number': container.truck_number,
        'loading_date': loading.loading_created_at.strftime('%d-%m-%Y') if loading else 'N/A',
        'seal_date': seal.seal_date.strftime('%d-%m-%Y') if seal else 'N/A',
        'seal_number': seal.seal_number if seal else 'N/A',
        'plant': contract.plant if contract else 'N/A',
        'product': contract.product if contract else 'N/A',
        'freight_forwarder': booking.freight_forwarder.name if booking and booking.freight_forwarder else 'N/A',
        'container_tare': container.container_tare,
        'weight_date': weight.weight_date.strftime('%d-%m-%Y') if weight else 'N/A',
        'weight_image_path': url_for('static', filename=f"images/weight/{weight.ws_image}", _external=True) if weight and weight.ws_image else 'N/A',
        'bags_type': container.bags_type if container else 'N/A',
        'loaded': loading.no_of_bags if loading else 'N/A',
        'gross': weight.gross_weight if weight else 'N/A',
        'tare': weight.tare_weight if weight else 'N/A',
        'net': weight.net_weight if weight else 'N/A',
        'ws_number': weight.ws_number if weight else 'N/A',
        'container_creator': container_creator_full_name,
        'loading_creator': loading_creator_full_name,
        'seal_creator': seal_creator_full_name,
        'weight_creator': weight_creator_full_name,
        'image_width': 610,
        'image_height': 350
    }

    html_content = render_template('container/container_pdf.html', **data)

    # Création du PDF
    pdf = HTML(string=html_content).write_pdf()

    # Création de la réponse HTTP avec le PDF
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=container_{container.container_name}.pdf'
    return response


@app.route('/report-booking')
@role_required('System Administrator', 'Finance Manager')
def report_booking():
    bookings = Booking.query.all()
    data = []
    for booking in bookings:
        contract = Contract.query.get(booking.contract_id)
        customer = Customer.query.get(contract.customer_id)
        data.append({
            'booking_id': booking.id,
            'booking_name': booking.booking_name,
            'plant': booking.plant,
            'product': booking.product,
            'contract_number': contract.contract_number,
            'customer_name': customer.name if customer else 'None',
            'customer_email': customer.email if customer else 'None',
            'quantity_planned': booking.quantity_planned,
            'quantity_loaded': booking.quantity_loaded
        })
    return render_template('booking/report_booking.html', data=data)


@app.route('/generate-excel-report/booking/<int:booking_id>', methods=['GET'])
@role_required('System Administrator', 'Logistics Manager', 'Logistics Data Analyst','Finance Manager')
def generate_excel_report_booking(booking_id):
    # Créez un objet BytesIO pour stocker le fichier Excel en mémoire
    output = BytesIO()

    # Récupérer les données pour la réservation spécifique
    booking = Booking.query.get(booking_id)
    if not booking:
        abort(404, description="Booking not found")

    # Récupérer le contrat associé
    contract = Contract.query.get(booking.contract_id)

    # Récupérer les conteneurs associés à la réservation
    containers = Container.query.filter_by(booking_id=booking_id).all()

    # Préparer les données pour le DataFrame
    data = []
    
    for container in containers:
        # Obtenez les données associées pour chaque conteneur
        weight = Weight.query.filter_by(container_id=container.id).first()
        seal = Seal.query.filter_by(container_id=container.id).first()
        loading = Loading.query.filter_by(container_id=container.id).first()
        
        # Ajouter les informations au DataFrame
        data.append({
            'Booking ID': booking.id,
            'Booking Name': booking.booking_name,
            'Contract Number': contract.contract_number if contract else 'None',
            'Container ID': container.id,
            'Container Name': container.container_name,
            'Arrival Date': container.arrival_date,
            'Truck Number': container.truck_number,
            'Plant': container.plant,
            'Product': container.product,
            'Freight Forwarder': container.freight_forwarder,
            'Container Tare': container.container_tare,
            'Bags Type': container.bags_type,
            'Net Weight': weight.net_weight if weight else 'None',
            'Weight Date': weight.weight_date if weight else 'None',
            'Seal Number': seal.seal_number if seal else 'None',
            'Seal Date': seal.seal_date if seal else 'None',
            'Loading No of Bags': loading.no_of_bags if loading else 'None',
            'Loading Labor': loading.labor if loading else 'None',
            'Loading Created At': loading.loading_created_at if loading else 'None',
        })
    
    # Convertir les données en DataFrame
    df = pd.DataFrame(data)
    
    # Créer un nouveau classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = 'Booking Report'
    
    # Écrire les en-têtes dans la première ligne avec mise en forme
    headers = df.columns.tolist()
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")  # Texte en gras et blanc
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Couleur d'arrière-plan bleu
        cell.alignment = Alignment(horizontal='center', vertical='center')  # Alignement centré

    # Écrire les données avec mise en forme
    for row_num, row_data in enumerate(df.values.tolist(), 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            cell.alignment = Alignment(horizontal='left', vertical='center')  # Alignement des cellules

    # Ajouter un tableau à la feuille de calcul
    table = Table(displayName="BookingTable", ref=f"A1:{ws.cell(row=1, column=len(headers)).column_letter}{len(df) + 1}")

    # Ajouter un style au tableau
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    table.tableStyleInfo = style

    # Ajouter le tableau à la feuille de calcul
    ws.add_table(table)

    # Ajuster la largeur des colonnes en fonction du contenu
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Obtenir le nom de la colonne
        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except:  # noqa: E722
                pass
        adjusted_width = max_length + 2  # Ajouter un peu d'espace pour la lisibilité
        ws.column_dimensions[column].width = adjusted_width

    # Enregistrer le fichier Excel en mémoire
    wb.save(output)
    output.seek(0)
    
    # Retourner le fichier Excel
    return send_file(output, as_attachment=True, download_name=f'Booking_{booking.booking_name}_Report.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')



@app.route('/generate_weight_images_pdf/<int:booking_id>')
@role_required('System Administrator','Finance Manager')
def generate_weight_images_pdf(booking_id):
    # Récupérer les données de la réservation
    booking = Booking.query.get(booking_id)
    if not booking:
        abort(404, description="Booking not found")



    if booking and booking.plant == 'BO':
        logo_filename = 'logo_bo.png'
    elif booking and booking.plant == 'BAB':
        logo_filename = 'logo_bab.png'
    else:
        logo_filename = 'default_logo.png'  # Un logo par défaut si nécessaire

    # Construire l'URL du logo
    logo_url = url_for('static', filename=logo_filename, _external=True)

    # Préparer les données pour le template

    # Récupérer les conteneurs associés à la réservation
    containers = Container.query.filter_by(booking_id=booking_id).all()

    # Récupérer les images des poids pour chaque conteneur
    weight_images = []
    for container in containers:
        weight = Weight.query.filter_by(container_id=container.id).first()
        if weight and weight.ws_image:
            weight_images.append({
                'ws_image': url_for('static', filename=f"images/weight/{weight.ws_image}", _external=True),
                'ws_number': weight.ws_number if weight.ws_number else 'N/A',
                'gross': weight.gross_weight,
                'tare': weight.tare_weight,
                'net': weight.net_weight,
                'container_name': container.container_name,
                'image_width': 610,
                'image_height': 420
            })

    if not weight_images:
        abort(404, description="No weight images found")

    # Préparer les données pour le template
    data = {
        'image_url': logo_url,
        'weight_images': weight_images,
        'booking_name': booking.booking_name,
        'plant': booking.plant,

    }

    # Génération du contenu HTML
    html_content = render_template('booking/weight_images_template.html', **data)

    # Création du PDF
    pdf = HTML(string=html_content).write_pdf()

    # Création de la réponse HTTP avec le PDF
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=weight_images_{booking.booking_name}.pdf'
    return response


# ---------------------------------------------Locales sales informations --------------------------------#
temp_truck_driver_data = None

@app.route('/localeSales/add_truck_driver', methods=['GET', 'POST'])
@login_required
@role_required('Logistics Manager', 'Warehouse Manager')
def add_truck_driver():
    global temp_truck_driver_data
    if request.method == 'POST':
        truck_number = request.form.get('truck_number')
        arrival_date = request.form.get('arrival_date')
        driver_name = request.form.get('driver_name')
        driver_phone_number = request.form.get('driver_phone_number')
        company = request.form.get('company')

        # Validation
        if not truck_number or not arrival_date or not driver_name or not company:
            flash('All fields are required!', 'error')
            return redirect(url_for('add_truck_driver'))
        
        if not validate_phone_number(driver_phone_number):
            flash('Invalid phone number!', 'error')
            return redirect(url_for('add_truck_driver'))

        temp_truck_driver_data = {
            'truck_number': truck_number,
            'arrival_date': arrival_date,
            'driver_name': driver_name,
            'driver_phone_number': driver_phone_number,
            'company': company,
            'created_by': current_user.username
        }
        return redirect(url_for('confirm_truck_driver_creation'))
    
    return render_template('localeSales/add_truck_driver.html')

@app.route('/localeSales/confirm_truck_driver_creation', methods=['GET', 'POST'])
@login_required
@role_required('Logistics Manager', 'Warehouse Manager')
def confirm_truck_driver_creation():
    global temp_truck_driver_data

    if request.method == 'POST':
        confirm_choice = request.form.get('choice')

        if confirm_choice == 'confirm':
            new_truck_driver = SalesTruckDriverInfo(
                truck_number=temp_truck_driver_data['truck_number'],
                arrival_date=datetime.strptime(temp_truck_driver_data['arrival_date'], '%Y-%m-%d'),
                driver_name=temp_truck_driver_data['driver_name'],
                driver_phone_number=temp_truck_driver_data['driver_phone_number'],
                company=temp_truck_driver_data['company'],
                created_by=current_user.username
            )

            db.session.add(new_truck_driver)
            db.session.commit()

            flash('Truck Driver added successfully', 'success')
        elif confirm_choice == 'cancel':
            flash('Truck Driver addition canceled.', 'info')

        temp_truck_driver_data = None
        return redirect(url_for('list_truck_drivers'))

    if temp_truck_driver_data:
        truck_driver_data = temp_truck_driver_data
    else:
        return redirect(url_for('add_truck_driver'))

    return render_template('localeSales/confirm_truck_driver_creation.html', truck_driver_data=truck_driver_data)

@app.route('/localeSales/truck_drivers/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('Logistics Manager', 'Warehouse Manager')
def edit_truck_driver(id):
    truck_driver = SalesTruckDriverInfo.query.get_or_404(id)

    if request.method == 'POST':
        truck_number = request.form['truck_number']
        arrival_date = request.form['arrival_date']
        driver_name = request.form['driver_name']
        driver_phone_number = request.form['driver_phone_number']
        company = request.form['company']

        # Validation
        if not truck_number or not arrival_date or not driver_name or not company:
            flash('All fields are required!', 'error')
            return redirect(url_for('edit_truck_driver', id=id))

        if not validate_phone_number(driver_phone_number):
            flash('Invalid phone number!', 'error')
            return redirect(url_for('edit_truck_driver', id=id))

        truck_driver.truck_number = truck_number
        truck_driver.arrival_date = datetime.strptime(arrival_date, '%Y-%m-%d')
        truck_driver.driver_name = driver_name
        truck_driver.driver_phone_number = driver_phone_number
        truck_driver.company = company
        truck_driver.created_by = current_user.username

        db.session.commit()
        flash('Truck Driver updated successfully', 'success')
        return redirect(url_for('list_truck_drivers'))

    return render_template('localeSales/edit_truck_driver.html', truck_driver=truck_driver)

@app.route('/localeSales/truck_driver/<int:id>/delete', methods=['POST'])
@login_required
@role_required('System Administrator')
def delete_truck_driver(id):
    truck_driver = SalesTruckDriverInfo.query.get_or_404(id)
    
    # Check if any related loading or weight info exists and handle accordingly
    if truck_driver.loading_info or truck_driver.weight_info:
        flash('Cannot delete truck driver with related loading or weight information!', 'error')
    else:
        db.session.delete(truck_driver)
        db.session.commit()
        flash('Truck Driver deleted successfully', 'success')

    return redirect(url_for('list_truck_drivers'))
 
@app.route('/localeSales/truck_drivers', methods=['GET'])
@login_required
@role_required('Logistics Manager', 'Warehouse Manager', 'Sales Manager','Finance Manager')
def list_truck_drivers():
    
    truck_drivers = SalesTruckDriverInfo.query.all()
    truck_drivers_data = []
    
    for driver in truck_drivers:
        loading_info = SalesLoadingInfo.query.filter_by(truck_id=driver.id).first()
        weight_info = SalesWeightInfo.query.filter_by(truck_id=driver.id).first()
        
        all_data_filled = loading_info is not None and weight_info is not None
        
        truck_drivers_data.append({
            'id': driver.id,
            'truck_number': driver.truck_number,
            'arrival_date': driver.arrival_date,
            'driver_name': driver.driver_name,
            'driver_phone_number': driver.driver_phone_number,
            'company': driver.company,
            'all_data_filled': all_data_filled
        })    # Vérifier si les tables liées sont complètes
 
    # Rendre le template avec les données nécessaires
    return render_template('localeSales/truck_drivers_list.html', truck_drivers=truck_drivers_data)

 





@app.route('/localeSales/truck_drivers/<int:id>/view', methods=['GET'])
@login_required
@role_required('Logistics Manager', 'Warehouse Manager', 'Sales Manager','Finance Manager')
def view_truck_driver(id):
    truck_driver = SalesTruckDriverInfo.query.get_or_404(id)

    # Récupérer les informations associées
    loading_info = SalesLoadingInfo.query.filter_by(truck_id=id).first()
    weight_info = SalesWeightInfo.query.filter_by(truck_id=id).first()
    
    # Passer les données au template pour affichage
    return render_template('localeSales/view_truck_driver.html',
                           truck_driver=truck_driver,
                           loading_info=loading_info,
                           weight_info=weight_info)
    
    
#---------Loading info---------


# Global variable
temp_loading_info_data = None

@app.route('/localeSales/add_loading_info/<int:truck_id>', methods=['GET', 'POST'])
@login_required
@role_required('Logistics Manager', 'Warehouse Manager')  # Restrict access based on roles
def add_loading_info(truck_id):
    truck = SalesTruckDriverInfo.query.get_or_404(truck_id)
    global temp_loading_info_data
    
    if request.method == 'POST':
        if request.form:
            bag_type = request.form.get('bag_type')
            no_of_bags = request.form.get('no_of_bags')
            labour_contract = request.form.get('labour_contract')
            loading_date = request.form.get('loading_date')
            destination = request.form.get('destination')

            # Validation
            if not bag_type or not no_of_bags or not labour_contract or not destination:
                flash('All fields are required!', 'error')
                return redirect(url_for('add_loading_info', truck_id=truck_id))

            temp_loading_info_data = {
                'truck_id': truck_id,
                'bag_type': bag_type,
                'no_of_bags': no_of_bags,
                'labour_contract': labour_contract,
                'loading_date': loading_date,
                'destination': destination,
                'created_by': current_user.username
            }
            return redirect(url_for('confirm_loading_info_creation'))

        elif 'confirm' in request.form:
            return redirect(url_for('confirm_loading_info_creation'))

    return render_template('localeSales/add_loading_info.html', truck_id=truck_id, truck=truck)

@app.route('/localeSales/confirm_loading_info_creation', methods=['GET', 'POST'])
@login_required
@role_required('Logistics Manager', 'Warehouse Manager')  # Restrict access based on roles
def confirm_loading_info_creation():
    global temp_loading_info_data

    if request.method == 'POST':
        confirm_choice = request.form.get('choice')

        if confirm_choice == 'confirm':
            new_loading_info = SalesLoadingInfo(
                truck_id=temp_loading_info_data['truck_id'],
                bag_type=temp_loading_info_data['bag_type'],
                no_of_bags=int(temp_loading_info_data['no_of_bags']),
                labour_contract=temp_loading_info_data['labour_contract'],
                loading_date=datetime.strptime(temp_loading_info_data['loading_date'], '%Y-%m-%d'),
                destination=temp_loading_info_data['destination'],
                created_by=current_user.username
            )

            db.session.add(new_loading_info)
            db.session.commit()

            flash('Loading Info added successfully', 'success')
        elif confirm_choice == 'cancel':
            flash('Loading Info addition canceled.', 'info')
        
        temp_loading_info_data = None
        return redirect(url_for('list_truck_drivers'))

    if temp_loading_info_data:
        loading_info_data = temp_loading_info_data
        truck = SalesTruckDriverInfo.query.get_or_404(loading_info_data['truck_id'])
    else:
        return redirect(url_for('add_loading_info', truck_id=temp_loading_info_data['truck_id']))

    return render_template('localeSales/confirm_loading_info_creation.html', loading_info_data=loading_info_data, truck=truck)

@app.route('/localeSales/loading_info/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('Logistics Manager', 'Warehouse Manager')  # Restrict access based on roles
def edit_loading_info(id):
    loading_info = SalesLoadingInfo.query.get_or_404(id)
    truck = SalesTruckDriverInfo.query.get_or_404(loading_info.truck_id)

    if request.method == 'POST':
        bag_type = request.form.get('bag_type')
        no_of_bags = request.form.get('no_of_bags')
        labour_contract = request.form.get('labour_contract')
        loading_date = request.form.get('loading_date')
        destination = request.form.get('destination')

        # Validation
        if not bag_type or not no_of_bags or not labour_contract or not destination:
            flash('All fields are required!', 'error')
            return redirect(url_for('edit_loading_info', id=id))

        loading_info.bag_type = bag_type
        loading_info.no_of_bags = int(no_of_bags)
        loading_info.labour_contract = labour_contract
        loading_info.loading_date = datetime.strptime(loading_date, '%Y-%m-%d')
        loading_info.destination = destination
        loading_info.created_by = current_user.username

        db.session.commit()
        flash('Loading Info updated successfully', 'success')
        return redirect(url_for('view_truck_driver', id=truck.id))

    return render_template('localeSales/edit_loading_info.html', loading_info=loading_info, truck=truck)

@app.route('/localeSales/loading_info/<int:id>/delete', methods=['POST'])
@login_required
@role_required('System Administrator')
def delete_loading_info(id):
    loading_info = SalesLoadingInfo.query.get_or_404(id)
    truck_id = loading_info.truck_id
    
    db.session.delete(loading_info)
    db.session.commit()
    flash('Loading Info deleted successfully', 'success')

    return redirect(url_for('view_truck_driver', id=truck_id))


# Global variable
temp_weight_info_data = None

@app.route('/localeSales/add_weight_info/<int:truck_id>', methods=['GET', 'POST'])
@login_required
@role_required('Logistics Manager', 'Warehouse Manager')
def add_weight_info(truck_id):
    truck = SalesTruckDriverInfo.query.get_or_404(truck_id)
    global temp_weight_info_data

    if request.method == 'POST':
        if 'preview' in request.form:
            gross_weight = request.form.get('gross_weight')
            tare_weight = request.form.get('tare_weight')
            net_weight = request.form.get('net_weight')
            ws_number = request.form.get('ws_number')
            weight_date = request.form.get('weight_date')
            ws_image = request.files.get('ws_image')  # Handle file uploads

            # Validation
            if not gross_weight or not tare_weight or not net_weight or not ws_number or not weight_date:
                flash('All fields are required!', 'error')
                return redirect(url_for('add_weight_info', truck_id=truck_id))

            image_filename = None
            if ws_image:
                image_filename = f"{ws_number}.jpg"
                ws_image.save(os.path.join(app.config['UPLOAD_FOLDER_WEIGHT_LOCALESALE'], image_filename))

            temp_weight_info_data = {
                'truck_id': truck_id,
                'gross_weight': gross_weight,
                'tare_weight': tare_weight,
                'net_weight': net_weight,
                'ws_number': ws_number,
                'ws_image': image_filename,
                'weight_date': weight_date,
                'created_by': current_user.username
            }
            return redirect(url_for('confirm_weight_info_creation'))

        elif 'confirm' in request.form:
            return redirect(url_for('confirm_weight_info_creation'))

    return render_template('localeSales/add_weight_info.html', truck=truck)

@app.route('/localeSales/confirm_weight_info_creation', methods=['GET', 'POST'])
@login_required
@role_required('Logistics Manager', 'Warehouse Manager')
def confirm_weight_info_creation():
    global temp_weight_info_data

    if request.method == 'POST':
        confirm_choice = request.form.get('choice')

        if confirm_choice == 'confirm':
            new_weight_info = SalesWeightInfo(
                truck_id=temp_weight_info_data['truck_id'],
                gross_weight=float(temp_weight_info_data['gross_weight']),
                tare_weight=float(temp_weight_info_data['tare_weight']),
                net_weight=float(temp_weight_info_data['net_weight']),
                ws_number=temp_weight_info_data['ws_number'],
                ws_image=temp_weight_info_data['ws_image'],
                weight_date=datetime.strptime(temp_weight_info_data['weight_date'], '%Y-%m-%d'),
                created_by=current_user.username
            )

            db.session.add(new_weight_info)
            db.session.commit()

            flash('Weight Info added successfully', 'success')
        elif confirm_choice == 'cancel':
            flash('Weight Info addition canceled.', 'info')

        temp_weight_info_data = None
        return redirect(url_for('list_truck_drivers'))

    if temp_weight_info_data:
        weight_info_data = temp_weight_info_data
        truck = SalesTruckDriverInfo.query.get_or_404(weight_info_data['truck_id'])
    else:
        return redirect(url_for('add_weight_info', truck_id=temp_weight_info_data['truck_id']))

    return render_template('localeSales/confirm_weight_info_creation.html', weight_info_data=weight_info_data, truck=truck)

@app.route('/localeSales/weight_info/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('Logistics Manager', 'System Administrator')
def edit_weight_info(id):
    weight_info = SalesWeightInfo.query.get_or_404(id)
    truck = SalesTruckDriverInfo.query.get_or_404(weight_info.truck_id)

    if request.method == 'POST':
        gross_weight = request.form.get('gross_weight')
        tare_weight = request.form.get('tare_weight')
        net_weight = request.form.get('net_weight')
        ws_number = request.form.get('ws_number')
        weight_date = request.form.get('weight_date')
        new_ws_image = request.files.get('ws_image')  # Handle file uploads

        # Validation
        if not gross_weight or not tare_weight or not net_weight or not ws_number or not weight_date:
            flash('All fields are required!', 'error')
            return redirect(url_for('edit_weight_info', id=id))

        image_filename = weight_info.ws_image
        if new_ws_image:
            image_filename = f"{ws_number}.jpg"
            new_ws_image.save(os.path.join(app.config['UPLOAD_FOLDER_WEIGHT_LOCALESALE'], image_filename))

        weight_info.gross_weight = float(gross_weight)
        weight_info.tare_weight = float(tare_weight)
        weight_info.net_weight = float(net_weight)
        weight_info.ws_number = ws_number
        weight_info.ws_image = image_filename
        weight_info.weight_date = datetime.strptime(weight_date, '%Y-%m-%d')
        weight_info.created_by = current_user.username

        db.session.commit()
        flash('Weight Info updated successfully', 'success')
        return redirect(url_for('view_truck_driver', id=truck.id))

    return render_template('localeSales/edit_weight_info.html', weight_info=weight_info, truck=truck)

@app.route('/localeSales/weight_info/<int:id>/delete', methods=['POST'])
@login_required
@role_required('System Administrator')
def delete_weight_info(id):
    weight_info = SalesWeightInfo.query.get_or_404(id)
    truck_id = weight_info.truck_id

    db.session.delete(weight_info)
    db.session.commit()
    flash('Weight Info deleted successfully', 'success')

    return redirect(url_for('view_truck_driver', id=truck_id))


@app.route('/localeSales/generate-sales-report', methods=['GET'])
@login_required
@role_required('Logistics Manager', 'Warehouse Manager','System Administrator','Finance Manager')  # Restrict access based on roles
def generate_sales_report():
    # Create a BytesIO object to store the Excel file in memory
    output = BytesIO()

    # Retrieve all truck information
    trucks = db.session.query(SalesTruckDriverInfo).all()
    
    # Prepare data for the DataFrame
    combined_data = []

    for truck in trucks:
        # Get loading and weight information for the current truck
        loading_info = truck.loading_info
        weight_info = truck.weight_info

        # Create a dictionary with all information in a single row
        row = {
            'Truck Number': truck.truck_number,
            'Arrival Date': truck.arrival_date,
            'Driver Name': truck.driver_name,
            'Driver Phone Number': truck.driver_phone_number,
            'Company': truck.company,
            'Created By': truck.created_by,
            'Created At': truck.created_at,
            
            # Loading Information
            'Loading Bag Type': loading_info.bag_type if loading_info else None,
            'Loading Number of Bags': loading_info.no_of_bags if loading_info else None,
            'Loading Labour Contract': loading_info.labour_contract if loading_info else None,
            'Loading Date': loading_info.loading_date if loading_info else None,
            'Loading Destination': loading_info.destination if loading_info else None,
            'Loading Created At': loading_info.created_at if loading_info else None,
            'Loading Created By': loading_info.created_by if loading_info else None,
            
            # Weight Information
            'Weight Gross Weight': weight_info.gross_weight if weight_info else None,
            'Weight Tare Weight': weight_info.tare_weight if weight_info else None,
            'Weight Net Weight': weight_info.net_weight if weight_info else None,
            'Weight WS Number': weight_info.ws_number if weight_info else None,
            'Weight Date': weight_info.weight_date if weight_info else None,
            'Weight Created At': weight_info.created_at if weight_info else None,
            'Weight Created By': weight_info.created_by if weight_info else None,
        }

        combined_data.append(row)

    # Convert data into DataFrame
    df = pd.DataFrame(combined_data)
    
    # Create an Excel file with a single sheet
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sales Report'

    # Add column headers with styling
    headers = df.columns.tolist()
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")  # Bold text and white color
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Blue background

    # Add data with styling
    for row_num, row_data in enumerate(df.values.tolist(), 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            cell.alignment = Alignment(horizontal='left', vertical='center')  # Cell alignment

    # Add a table to the worksheet
    table = Table(displayName="SalesReport", ref=f"A1:{ws.cell(row=1, column=len(headers)).column_letter}{len(df) + 1}")

    # Add a style to the table
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    table.tableStyleInfo = style

    # Add the table to the worksheet
    ws.add_table(table)

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:  # noqa: E722
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the Excel file to the BytesIO object
    wb.save(output)
    output.seek(0)
    
    # Define the filename with the current date and time
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f'BAB_Sales_Report_{now}.xlsx'
    
    # Return the Excel file
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')



@app.route('/generate_local_sales_pdf/<int:truck_id>')
@role_required('System Administrator', 'Logistics Manager', 'Logistics Data Analyst','Finance Manager')
def generate_local_sales_pdf(truck_id):
    # Récupérer les données du camion
    truck = SalesTruckDriverInfo.query.get(truck_id)
    if not truck:
        abort(404, description="Truck not found")

    # Récupérer les données associées
    loading = SalesLoadingInfo.query.filter_by(truck_id=truck_id).first()
    weight = SalesWeightInfo.query.filter_by(truck_id=truck_id).first()
    
    created_by_user = User.query.filter_by(username=truck.created_by).first()
    created_by_full_name = created_by_user.full_name if created_by_user else 'N/A'
    
    # Choisir le logo (vous pouvez personnaliser la logique selon vos besoins)
    logo_filename = 'logo_bab.png'  # Un logo par défaut si nécessaire
    logo_url = url_for('static', filename=logo_filename, _external=True)

    # Préparer les données pour le template
    data = {
        'image_url': logo_url,
        'truck_number': truck.truck_number,
        'arrival_date': truck.arrival_date.strftime('%d-%m-%Y'),
        'driver_name': truck.driver_name,
        'driver_phone_number': truck.driver_phone_number,
        'company': truck.company,
        'loading_date': loading.loading_date.strftime('%d-%m-%Y') if loading else 'N/A',
        'bag_type': loading.bag_type if loading else 'N/A',
        'no_of_bags': loading.no_of_bags if loading else 'N/A',
        'gross_weight': weight.gross_weight if weight else 'N/A',
        'tare_weight': weight.tare_weight if weight else 'N/A',
        'net_weight': weight.net_weight if weight else 'N/A',
        'ws_number': weight.ws_number if weight else 'N/A',
        'ws_image_path': url_for('static', filename=f"images/weight_localSales/{weight.ws_image}", _external=True)  if weight and weight.ws_image else 'N/A',
        'created_at': truck.created_at.strftime('%d-%m-%Y %H:%M:%S'),
        'created_by': truck.created_by,
        'created_by_full_name': created_by_full_name,
        'image_width': 610,
        'image_height': 420
    }

    # Rendre le contenu HTML du PDF
    html_content = render_template('localeSales/local_sales_pdf.html', **data)

    # Création du PDF avec WeasyPrint
    pdf = HTML(string=html_content).write_pdf()

    # Création de la réponse HTTP avec le PDF
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=local_sales_{truck.truck_number}.pdf'
    return response


@app.route('/admin', methods=['GET'])
@role_required('System Administrator')
@login_required
def admin_interface():
    """
    Display the admin interface with various links for administrative actions.
    Accessible only to authenticated users with the 'admin' role.
    """
    # Retrieve user actions for display (if needed)
    
    # Render the HTML template for the admin interface
    return render_template('admin/admin_interface.html')



if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True)
